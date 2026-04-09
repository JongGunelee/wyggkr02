"""
================================================================================
 [초고속 텍스트 탐색기 (Search Two Items) v34.1.16] (Ultimate Quantum)
================================================================================
- 아키텍처: Clean Layer Architecture (Domain / Presentation / Application)
- 주요 기능: 수천 개 엑셀 파일 내 특정 키워드(AND/OR) 초고속 탐색 및 구조 분석
- 가이드라인 준수: 00 PRD 가이드.md | AI_CODING_GUIDELINES_2026.md
- 무결성 보증: Read-Only 모드 탐색으로 데이터 변종 방지
================================================================================
"""
import sys
try:
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8')
except: pass
import os
import openpyxl
import time
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import threading

# ═══════════════════════════════════════════════════════════
# LAYER 1: DOMAIN & INFRASTRUCTURE (Core Search Engine)
# ═══════════════════════════════════════════════════════════
class SearchEngine:
    @staticmethod
    def analyze_excel_structure(file_path):
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet_names = wb.sheetnames
        ws = wb.active
        headers = []
        for r in range(1, 11):
            row_data = []
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                val = cell.value
                if val is None:
                    for rg in ws.merged_cells.ranges:
                        if cell.coordinate in rg:
                            val = ws.cell(row=rg.min_row, column=rg.min_col).value; break
                txt = str(val or "").strip()
                if txt and len(txt) < 25: row_data.append(txt)
            if len(set(row_data)) >= 2:
                headers = list(dict.fromkeys(row_data)); break
        wb.close()
        return {"sheets": sheet_names, "headers": headers}

    @staticmethod
    def perform_search(config, update_callback):
        found_cnt = 0
        seen_data = set()
        files = []
        for root_dir, _, fs in os.walk(config['base_path']):
            for f in fs:
                if config['file_kw'] in f and f.lower().endswith((".xlsx", ".xlsm")):
                    files.append(os.path.join(root_dir, f))
        files.sort(key=os.path.getmtime, reverse=True)

        for f_path in files:
            if found_cnt >= config['limit']: break
            update_callback("status", f"진행: {os.path.basename(f_path)}")
            try:
                wb = openpyxl.load_workbook(f_path, data_only=True, read_only=True)
                if config['sheet_scope'] == "첫 번째 시트": target_ws = [wb.worksheets[0]]
                elif config['sheet_scope'] == "전체 워크시트": target_ws = wb.worksheets
                else: target_ws = [s for s in wb.worksheets if s.title == config['sheet_scope']] or wb.worksheets
                
                for ws in target_ws:
                    if found_cnt >= config['limit']: break
                    h1_idx, h2_idx, start_row = -1, -1, 1
                    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), 1):
                        r_s = [str(c or "").strip() for c in row]
                        for i, cell_txt in enumerate(r_s):
                            if config['h1_name'] and config['h1_name'] in cell_txt:
                                if h1_idx == -1: h1_idx = i
                            if config['h2_name'] and config['h2_name'] in cell_txt:
                                if h2_idx == -1: h2_idx = i
                        if h1_idx != -1 or (config['g_mode'] != "모든 열 (전역 검색)" and h2_idx != -1):
                            start_row = r_idx + 1; break
                    
                    for row in ws.iter_rows(min_row=start_row, values_only=True):
                        if found_cnt >= config['limit']: break
                        row_s = [str(c or "").strip() for c in row]
                        if not any(row_s): continue
                        
                        match1 = any(k in row_s[h1_idx] for k in config['kw1']) if h1_idx != -1 and config['kw1'] else (not config['kw1'])
                        match2 = any(k in " ".join(row_s) for k in config['kw2']) if config['g_mode'] == "모든 열 (전역 검색)" and config['kw2'] \
                             else (any(k in row_s[h2_idx] for k in config['kw2']) if h2_idx != -1 and config['kw2'] else (not config['kw2']))
                        
                        if match1 and match2:
                            v1 = row_s[h1_idx] if h1_idx != -1 else ""
                            v2 = row_s[h2_idx] if h2_idx != -1 else "/".join(row_s)
                            if config['unique'] and (v1, v2) in seen_data: continue
                            seen_data.add((v1, v2))
                            found_cnt += 1
                            update_callback("result", {"cnt": found_cnt, "file": os.path.basename(f_path), "sheet": ws.title, 
                                                       "h1": config['h1_name'], "v1": v1, "h2": config['h2_name'], 
                                                       "v2": v2, "full": " / ".join(row_s), "g_mode": config['g_mode']})
                wb.close()
            except: pass
        return found_cnt

# ═══════════════════════════════════════════════════════════
# LAYER 2: PRESENTATION (Guide-Integrated View)
# ═══════════════════════════════════════════════════════════
class GuideView:
    def __init__(self, root, controller):
        self.root = root
        self.controller = controller
        self.root.title("Excel 로컬 지능형 탐색기 (v34.1.16)")
        self.root.geometry("850x700")
        self._build_ui()

    def _build_ui(self):
        tk.Label(self.root, text="Excel 지능형 로컬 통합 탐색 시스템", font=("Malgun Gothic", 14, "bold"), pady=8, fg="#1A237E").pack()

        main = tk.Frame(self.root, padx=15)
        main.pack(fill="both", expand=True)

        # 1. 경로 설정부
        path_f = tk.Frame(main, pady=2)
        path_f.pack(fill="x")
        tk.Label(path_f, text="[DIR] 경로:", width=6, anchor="w", font=("Malgun Gothic", 9)).pack(side="left")
        tk.Entry(path_f, textvariable=self.controller.path_var, font=("Malgun Gothic", 9)).pack(side="left", fill="x", expand=True, padx=5)
        tk.Button(path_f, text="폴더 선택", command=self.controller.handle_folder_select, font=("Malgun Gothic", 8)).pack(side="left", padx=1)
        tk.Button(path_f, text="[SCAN] 구조 분석", bg="#E3F2FD", command=self.controller.handle_structure_analysis, font=("Malgun Gothic", 8, "bold")).pack(side="left", padx=1)

        # 2. 파일 필터부
        opt_f = tk.Frame(main, pady=2)
        opt_f.pack(fill="x")
        tk.Label(opt_f, text="[FILE] 파일:", width=6, anchor="w", font=("Malgun Gothic", 9)).pack(side="left")
        tk.Entry(opt_f, textvariable=self.controller.file_keyword_var, font=("Malgun Gothic", 9), width=35).pack(side="left", padx=5)
        tk.Button(opt_f, text="파일 선택(패턴추출)", command=self.controller.handle_file_keyword_pick, font=("Malgun Gothic", 8)).pack(side="left", padx=2)
        
        tk.Label(opt_f, text="| 🔢 개수:", font=("Malgun Gothic", 9)).pack(side="left", padx=(10,0))
        tk.Entry(opt_f, textvariable=self.controller.max_findings_var, width=5, font=("Malgun Gothic", 9)).pack(side="left", padx=5)
        tk.Checkbutton(opt_f, text="중복제거", variable=self.controller.unique_check_var, font=("Malgun Gothic", 8)).pack(side="left", padx=5)

        # 3. 검색 매칭 설정
        cfg_f = tk.Frame(main, pady=8, relief="groove", bd=1)
        cfg_f.pack(fill="x")

        sel_f = tk.Frame(cfg_f, pady=2)
        sel_f.pack(fill="x", padx=5)
        tk.Label(sel_f, text="[LOC] 열 헤더 1:", font=("Malgun Gothic", 8, "bold"), fg="#555").pack(side="left")
        self.cb1 = ttk.Combobox(sel_f, textvariable=self.controller.header1_var, font=("Malgun Gothic", 9), width=15)
        self.cb1.pack(side="left", padx=(2, 10))
        tk.Label(sel_f, text="[DOC] 대상 시트:", font=("Malgun Gothic", 8, "bold"), fg="#555").pack(side="left")
        self.s_cb = ttk.Combobox(sel_f, textvariable=self.controller.sheet_scope_var, font=("Malgun Gothic", 9), width=15)
        self.s_cb.pack(side="left", padx=(2, 10))
        tk.Label(sel_f, text="[FIND] 범위:", font=("Malgun Gothic", 8, "bold"), fg="#555").pack(side="left")
        self.g_cb = ttk.Combobox(sel_f, textvariable=self.controller.search_scope2_var, values=["모든 열 (전역)", "특정 열 선택"], font=("Malgun Gothic", 9), width=14)
        self.g_cb.pack(side="left", padx=(2, 5))
        self.g_cb.bind("<<ComboboxSelected>>", self.controller.handle_scope_toggle)
        self.cb2 = ttk.Combobox(sel_f, textvariable=self.controller.header2_var, font=("Malgun Gothic", 9), width=12, state="disabled")
        self.cb2.pack(side="left", padx=2)

        kw_f = tk.Frame(cfg_f, pady=4)
        kw_f.pack(fill="x", padx=5)
        k1_wrap = tk.Frame(kw_f); k1_wrap.pack(side="left", fill="x", expand=True, padx=2)
        tk.Label(k1_wrap, text=" ↓ [열 1] 찾을 단어들(OR):", font=("Malgun Gothic", 7), fg="blue").pack(anchor="w")
        self.kw1 = tk.Text(k1_wrap, height=3, font=("Malgun Gothic", 9)); self.kw1.pack(fill="x")
        self.kw1.insert("1.0", "베란다 누수보수(전후면)")
        
        k2_wrap = tk.Frame(kw_f); k2_wrap.pack(side="left", fill="x", expand=True, padx=2)
        tk.Label(k2_wrap, text=" ↓ [범위] 찾을 단어들(OR/AND):", font=("Malgun Gothic", 7), fg="blue").pack(anchor="w")
        self.kw2 = tk.Text(k2_wrap, height=3, font=("Malgun Gothic", 9)); self.kw2.pack(fill="x")

        # 4. 결과창
        res_f = tk.Frame(main, pady=5)
        res_f.pack(fill="both", expand=True)
        self.res_txt = tk.Text(res_f, font=("Malgun Gothic", 9), bg="#fafafa", state="disabled", wrap="word") 
        scr = tk.Scrollbar(res_f, command=self.res_txt.yview)
        self.res_txt.config(yscrollcommand=scr.set); scr.pack(side="right", fill="y")
        self.res_txt.pack(side="left", fill="both", expand=True)

        # 5. 액션 버튼부
        act_f = tk.Frame(main, pady=10)
        act_f.pack(fill="x")
        self.run_btn = tk.Button(act_f, text="[START] 탐색 시작", font=("Malgun Gothic", 10, "bold"), bg="#2E7D32", fg="white", width=25, command=self.controller.handle_search_start)
        self.run_btn.pack(side="left", expand=True, padx=5)
        tk.Button(act_f, text="[COPY] 복사", font=("Malgun Gothic", 10, "bold"), bg="#1565C0", fg="white", width=20, command=self.controller.handle_copy).pack(side="left", expand=True, padx=5)
        tk.Button(act_f, text="📖 사용 방법", font=("Malgun Gothic", 9), bg="#607D8B", fg="white", command=self.controller.show_help_dialog).pack(side="left", padx=10)

        # 6. 하단 상태바
        self.stat = tk.Label(self.root, text="준비됨", bd=1, relief="sunken", anchor="w", font=("Malgun Gothic", 8), padx=10)
        self.stat.pack(side="bottom", fill="x")

# ═══════════════════════════════════════════════════════════
# LAYER 3: APPLICATION (App Controller)
# ═══════════════════════════════════════════════════════════
class AppController:
    def __init__(self, root):
        self.root = root
        self.engine = SearchEngine()
        # [이식성 강화] 하드코딩된 절대 경로를 현재 작업 디렉토리로 초기화
        self.path_var = tk.StringVar(value=os.getcwd())
        self.file_keyword_var = tk.StringVar(value="발송_솔루션_목록_(기숙사 및 사택)")
        self.max_findings_var = tk.StringVar(value="10")
        self.header1_var = tk.StringVar(value="계약명")
        self.header2_var = tk.StringVar(value="공사내용")
        self.sheet_scope_var = tk.StringVar(value="전체 워크시트")
        self.search_scope2_var = tk.StringVar(value="모든 열 (전역)")
        self.unique_check_var = tk.BooleanVar(value=True)
        self.is_searching = False
        self.view = GuideView(root, self)
        
        default_h = ["계약명", "공사내용", "프로젝트명", "비고"]
        self.view.cb1['values'] = default_h
        self.view.cb2['values'] = default_h
        self.view.s_cb['values'] = ["전체 워크시트", "첫 번째 시트"]

    def handle_folder_select(self):
        p = filedialog.askdirectory(initialdir=self.path_var.get())
        if p: self.path_var.set(p)

    def handle_file_keyword_pick(self):
        f = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xlsm")], initialdir=self.path_var.get())
        if f:
            self.file_keyword_var.set(os.path.splitext(os.path.basename(f))[0])

    def handle_scope_toggle(self, _):
        self.view.cb2.config(state="normal" if "특정" in self.search_scope2_var.get() else "disabled")

    def handle_structure_analysis(self):
        f = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xlsm")], initialdir=self.path_var.get())
        if not f: return
        self.view.stat.config(text="구조 분석 중...")
        def run():
            try:
                data = self.engine.analyze_excel_structure(f)
                self.root.after(0, lambda: self._update_ui(data))
            except Exception as e: self.root.after(0, lambda: messagebox.showerror("오류", str(e)))
        threading.Thread(target=run, daemon=True).start()

    def _update_ui(self, data):
        self.view.s_cb['values'] = ["전체 워크시트", "첫 번째 시트"] + data['sheets']
        sh = sorted(data['headers'])
        self.view.cb1['values'] = sh; self.view.cb2['values'] = sh
        if sh: self.header1_var.set(sh[0])
        if len(sh)>1: self.header2_var.set(sh[1])
        self.view.stat.config(text="구조 분석 완료")

    def handle_search_start(self):
        if self.is_searching: return
        k1 = [l.strip() for l in self.view.kw1.get("1.0", "end").split('\n') if l.strip()]
        k2 = [l.strip() for l in self.view.kw2.get("1.0", "end").split('\n') if l.strip()]
        if not k1 and not k2: return messagebox.showwarning("입력 부족", "찾을 단어를 입력하세요.")
        
        try: lim = int(self.max_findings_var.get())
        except: lim = 10
        self.is_searching = True
        self.view.run_btn.config(state="disabled", text="⌛ 검색 중...")
        self.view.res_txt.config(state="normal")
        self.view.res_txt.delete("1.0", "end")
        self.view.res_txt.insert("end", f"[START] 탐색 가동 ({time.strftime('%H:%M:%S')})\n{'-'*80}\n")
        self.view.res_txt.config(state="disabled")

        conf = {"base_path": self.path_var.get().strip(), "file_kw": self.file_keyword_var.get().strip(),
                "limit": lim, "kw1": k1, "kw2": k2, "h1_name": self.header1_var.get(), 
                "h2_name": self.header2_var.get(), "sheet_scope": self.sheet_scope_var.get(),
                "g_mode": "모든 열 (전역 검색)" if "전역" in self.search_scope2_var.get() else "특정 열 선택", "unique": self.unique_check_var.get()}

        threading.Thread(target=lambda: self._run_engine(conf), daemon=True).start()

    def _run_engine(self, conf):
        cnt = self.engine.perform_search(conf, self._cb)
        self.root.after(0, lambda: self._done(cnt))

    def _cb(self, type, d):
        if type == "status": self.root.after(0, lambda: self.view.stat.config(text=d))
        elif type == "result": self.root.after(0, lambda: self._add_res_ui(d))

    def _add_res_ui(self, d):
        self.view.res_txt.config(state="normal")
        self.view.res_txt.insert("end", f"[OK] #{d['cnt']} | {d['file']} | {d['sheet']}\n")
        self.view.res_txt.insert("end", f" 🔹 {d['h1']}: {d['v1']}\n")
        # [핵심 수저] 말줄임표(slice) 제거하고 전체 내용 출력
        v_sub = f" 🔸 {d['h2']}: {d['v2']}\n" if "특정" in d['g_mode'] else f" 🔸 요약: {d['full']}\n"
        self.view.res_txt.insert("end", v_sub + f"{'-'*80}\n")
        self.view.res_txt.see("end"); self.view.res_txt.config(state="disabled")

    def _done(self, c):
        self.is_searching = False
        self.view.run_btn.config(state="normal", text="🚀 탐색 시작")
        self.view.stat.config(text=f"완료: {c}건")
        messagebox.showinfo("탐색 완료", f"총 {c}건 발견.")

    def handle_copy(self):
        c = self.view.res_txt.get("1.0", "end-1c")
        if not c.strip(): return
        self.root.clipboard_clear(); self.root.clipboard_append(c)
        messagebox.showinfo("복사", "결과가 복사되었습니다.")

    def show_help_dialog(self):
        help_msg = (
            "📖 [Excel 지능형 로컬 탐색기] 사용 가이드\n\n"
            "1. 폴더 선택: 검색할 파일들이 있는 폴더를 고르세요.\n"
            "2. ✨ 구조 분석: 폴더 내 파일 하나를 선택해 '시트'와 '헤더'를 학습시킵니다.\n"
            "3. 📝 파일 선택: 찾을 파일의 이름 규칙을 자동으로 입력합니다.\n"
            "4. 단어 입력: \n"
            "   - 줄바꿈으로 입력 시 각 단어 중 하나라도 있으면 매칭(OR)\n"
            "   - 열 1과 범위 양쪽에 입력 시 두 조건 모두 만족(AND)\n"
            "5. 범위 설정: '전역'은 전체 데이터 조사, '특정 열'은 지정 열만 조사\n"
            "6. 🚀 탐색 시작: 연산 후 결과를 확인하고 📋 복사하세요."
        )
        messagebox.showinfo("사용 방법", help_msg)

if __name__ == "__main__":
    root = tk.Tk()

    # [v34.1.21] Stealth Launch 대응: 창을 최상단으로 강제 부각
    root.lift()
    root.attributes('-topmost', True)
    root.after(100, lambda: root.attributes('-topmost', False))

    AppController(root)
    root.mainloop()
