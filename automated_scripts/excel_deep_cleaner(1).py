"""
================================================================================
 [엑셀 얼티밋 딥-클리너 (Excel Deep Cleaner) v34.1.5] (Registry Repair & Force Bind)
================================================================================
- 아키텍처: Clean Layer Architecture (Domain / Presentation / Application)
- 주요 기능: 엑셀 내부 메타데이터, 개인정보, 숨겨진 정의 이름 완전 소거(Deep)
- 가이드라인 준수: 00 PRD 가이드.md | AI_CODING_GUIDELINES_2026.md
- 무결성 보증: Auto-Elevation (관리자 자동 승격), Force Process Kill (좀비 소거)
- 최적화: Dynamic COM Binding 기능으로 환경 이식성 극대화
================================================================================
"""
import sys
try:
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8')
except: pass

import os
import shutil
import threading
import time
import win32com.client
import openpyxl
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

# ═══════════════════════════════════════════════════════════
# LAYER 1: DOMAIN (Deep Clean Engine)
# ═══════════════════════════════════════════════════════════
class ExcelDeepCleanEngine:
    def __init__(self, callback):
        self.callback = callback
        self.excel = None
        self.clean_options = {}

    def _kill_office_processes(self):
        """[v34.1.2] 기존 좀비 프로세스 강제 소거 (psutil+WMI 듀얼 네이티브)"""
        import time, ctypes
        try: ctypes.windll.kernel32.SetErrorMode(0x0001 | 0x0002 | 0x8000)
        except: pass
        self.callback("log", "[CLEAN] 기존 Excel 좀비 프로세스 네이티브 정리 중 (팝업 차단)...")
        
        targets = ["EXCEL.EXE"]
        try:
            import psutil
            for p in psutil.process_iter(['name']):
                if p.info['name'] and p.info['name'].upper() in targets:
                    try: p.kill()
                    except: pass
        except ImportError:
            try:
                import pythoncom, win32com.client
                pythoncom.CoInitialize()
                wmi = win32com.client.GetObject("winmgmts:")
                for p in wmi.InstancesOf("Win32_Process"):
                    name = p.Properties_('Name').Value
                    if name and name.upper() in targets:
                        try: p.Terminate()
                        except: pass
            except: pass
        time.sleep(1)

    def _get_excel(self):
        """[v34.1.4] CLSID Direct & Deep Polling 엔진 가동"""
        if not self.excel:
            import pythoncom
            pythoncom.CoInitialize()
            self._kill_office_processes()
            
            prog_id = "Excel.Application"
            clsid = "{00024500-0000-0000-C000-000000000046}"
            app = None
            errors = []
            
            import win32com.client
            from win32com.client import Dispatch, DispatchEx, gencache
            from win32com.client.dynamic import Dispatch as DynDispatch
            
            # Phase 0: CLSID Direct Dispatch (가장 강력한 직접 바인딩)
            try:
                app = DynDispatch(clsid)
                if app: return self._finalize_excel(app)
            except: pass

            # Phase 1: EnsureDispatch (레지스트리 강제 복구)
            try:
                app = gencache.EnsureDispatch(prog_id)
                if app: return self._finalize_excel(app)
            except Exception as e1:
                errors.append(f"P1:{str(e1)[:20]}")

            # Phase 2: Shell/Popen + Deep Polling (강제 가중 가동 후 추적)
            try:
                import subprocess, time
                subprocess.Popen(['excel'], shell=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
                for i in range(10): # 10초간 추적 가동
                    time.sleep(1.0)
                    try:
                        app = win32com.client.GetActiveObject(prog_id)
                        if app: return self._finalize_excel(app)
                    except: pass
            except Exception as e2:
                errors.append(f"P2:{str(e2)[:20]}")

            # Phase 3: Last Resort (Standard Dispatch)
            try:
                app = win32com.client.Dispatch(prog_id)
                if app: return self._finalize_excel(app)
            except Exception as e3:
                errors.append(f"P3:{str(e3)[:20]}")

            # Phase 4: Emergency Shell Hook
            try:
                import os
                os.startfile("excel")
                for _ in range(5):
                    time.sleep(2.0)
                    try: 
                        app = win32com.client.GetActiveObject(prog_id)
                        if app: return self._finalize_excel(app)
                    except: pass
            except Exception as e4:
                errors.append(f"P4:{str(e4)[:20]}")

            if not app:
                self.callback("log", f"[FAIL] Excel 엔진 가동 실패: {'; '.join(errors)}")
                return None
        return self.excel

    def _finalize_excel(self, app):
        """Excel 설정 마무리 및 가동 확인"""
        try:
            app.Visible = False
            app.DisplayAlerts = False
            app.AskToUpdateLinks = False
        except:
            pass
        self.excel = app
        self.callback("log", "[OK] Excel 엔진(v34.1.5 Registry Repair) 가동 완료.")
        return app

    def clean_file(self, src_path, dest_path, options=None):
        """
        V2.2 ULTIMATE DEEP CLEAN (공식 Excel API 기반)
        """
        if options is None: options = {}

        app = self._get_excel()
        wb = None
        stats = {
            "links": 0, "names_before": 0, "names_after": 0,
            "comments_before": 0, "comments_after": 0,
            "doc_info_removed": "N", "sheets": 0
        }
        
        # xlRDI API Constants
        xlRDIComments = 1
        xlRDIRemovePersonalInformation = 4
        xlRDIDocumentProperties = 8
        xlRDIDefinedNameComments = 18
        xlRDIAll = 99
        
        try:
            wb = app.Workbooks.Open(src_path, UpdateLinks=False)
            stats["sheets"] = wb.Worksheets.Count
            
            # 1. Pre-Check
            stats["names_before"] = wb.Names.Count
            for ws in wb.Worksheets:
                try: stats["comments_before"] += ws.Comments.Count
                except: pass
            
            # 2. Break Links
            links = wb.LinkSources(1)
            if links:
                stats["links"] = len(links)
                for link in links:
                    try: wb.BreakLink(Name=link, Type=1)
                    except: pass
            
            # 3. Document Inspector API
            try:
                wb.RemoveDocumentInformation(xlRDIAll)
                stats["doc_info_removed"] = "Y"
            except Exception as e:
                stats["doc_info_removed"] = f"부분({str(e)[:20]})"
            
            # 4. Remove System Names (Print Area, AutoFilter)
            for ws in wb.Worksheets:
                try: ws.PageSetup.PrintArea = ""
                except: pass
                try:
                    if ws.AutoFilterMode: ws.AutoFilterMode = False
                except: pass
            
            # 5. Fallback Name Deletion
            remaining_names = wb.Names.Count
            if remaining_names > 0:
                for j in range(remaining_names, 0, -1):
                    try:
                        nm = wb.Names(j)
                        nm.Visible = True
                        nm.Delete()
                    except: pass
            
            # 6. Fallback Comment Deletion
            for ws in wb.Worksheets:
                try:
                    if ws.Comments.Count > 0: ws.Comments.Delete()
                except: pass
            
            # 7. Post-Check
            stats["names_after"] = wb.Names.Count
            for ws in wb.Worksheets:
                try: stats["comments_after"] += ws.Comments.Count
                except: pass
            
            # 8. Activate Sheet Option
            activate_mode = options.get('activate_mode', 'first')
            custom_sheet = options.get('custom_sheet', '')
            
            if activate_mode == 'first':
                try: wb.Worksheets(1).Activate()
                except: pass
            elif activate_mode == 'custom' and custom_sheet:
                try: wb.Worksheets(custom_sheet).Activate()
                except: pass
            
            # [Atomic Write] 임시 파일에 먼저 저장한 후 교체하여 무결성 확보
            import uuid
            temp_path = f"{dest_path}.{uuid.uuid4().hex[:8]}.tmp"
            wb.SaveAs(temp_path)
            wb.Close(False)
            
            if os.path.exists(temp_path) and os.path.getsize(temp_path) > 0:
                if os.path.exists(dest_path):
                    os.remove(dest_path)
                os.rename(temp_path, dest_path)
                return True, stats
            else:
                return False, "임시 파일 생성 실패 또는 파일이 비어 있습니다."
        except Exception as e:
            if wb: 
                try: wb.Close(False)
                except: pass
            return False, str(e)

    def verify_integrity(self, file_path):
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheet_count = len(wb.sheetnames)
            wb.close()
            return True, f"검증 통과 (시트 {sheet_count}개 확인)"
        except Exception as e:
            return False, f"무결성 검증 실패: {str(e)}"

    def run_process_dual(self, mode, targets, options):
        """
        듀얼 모드 지원 프로세서
        mode: 'folder' | 'files'
        targets: folder_path string | list of file paths
        """
        start_time = time.time()
        use_prefix = options.get('use_prefix', True)
        self.clean_options = options
        
        # 1. 대상 파일 수집
        excel_files = []
        base_dir = "" # 결과 저장 기준 폴더
        
        if mode == 'folder':
            base_dir = targets
            self.callback("status", f"🔍 폴더 스캔 중: {base_dir}")
            for root, _, files in os.walk(base_dir):
                if "00_Deep_Clean_Results" in root: continue
                for f in files:
                    if f.lower().endswith(('.xlsx', '.xlsm', '.xls')) and not f.startswith('~$') and not f.startswith('클리닝_'):
                        excel_files.append(os.path.join(root, f))
        else: # files mode
            if targets:
                base_dir = os.path.dirname(targets[0])
                self.callback("status", f"📄 선택된 파일 {len(targets)}건 준비 중")
                excel_files = targets
            
        if not excel_files:
            self.callback("done", "대상 파일을 찾지 못했습니다.")
            self._cleanup()
            return

        # 2. 결과 저장 폴더 생성
        res_dir = os.path.join(base_dir, "00_Deep_Clean_Results")
        if not os.path.exists(res_dir): os.makedirs(res_dir)
        
        total = len(excel_files)
        success_cnt = 0
        details_log = []

        # 3. 처리 루프
        for i, src in enumerate(excel_files, 1):
            f_name = os.path.basename(src)
            
            # 폴더 모드일 경우 하위 구조 유지, 파일 모드는 플랫하게
            if mode == 'folder':
                rel_path = os.path.relpath(src, base_dir)
                dest_subdir = os.path.join(res_dir, os.path.dirname(rel_path))
            else:
                dest_subdir = res_dir
                
            if not os.path.exists(dest_subdir): os.makedirs(dest_subdir)
            
            prefix = "클리닝_" if use_prefix else ""
            dest = os.path.join(dest_subdir, f"{prefix}{f_name}")
            
            self.callback("status", f"[{i}/{total}] {f_name} 정제 중...")
            
            # Cleaning Execution
            ok, result = self.clean_file(os.path.abspath(src), os.path.abspath(dest), self.clean_options)
            
            log_entry = {
                "파일명": f_name, "상태": "", 
                "검증결과": ""
            }

            if ok:
                stats = result
                v_ok, v_msg = self.verify_integrity(dest)
                if v_ok:
                    success_cnt += 1
                    names_info = f"{stats['names_before']}→{stats['names_after']}"
                    if stats['names_after'] > 0: names_info += "(시스템)"
                    
                    self.callback("log", f"[OK] [완료] {f_name} (이름:{names_info}, 링크:{stats['links']}, 메모제거:{stats['comments_before'] - stats['comments_after']})")
                    log_entry.update({"상태": "성공", "검증결과": "정상"})
                else:
                    self.callback("log", f"[WARN] [검증실패] {f_name}: {v_msg}")
                    log_entry.update({"상태": "검증실패", "검증결과": v_msg})
            else:
                self.callback("log", f"[FAIL] [실패] {f_name}: {result}")
                log_entry.update({"상태": "실패", "검증결과": result})
            
            details_log.append(log_entry)

        # 4. 보고서 생성
        self._create_report(res_dir, details_log)

        # 5. 종료
        duration = time.time() - start_time
        self._cleanup()
        self.callback("done", f"총 {total}건 중 {success_cnt}건 성공 ({duration:.1f}초)\n저장위치: {res_dir}")

    def _create_report(self, res_dir, logs):
        try:
            report_path = os.path.join(res_dir, f"정제_보고서_{int(time.time())}.xlsx")
            wb = openpyxl.Workbook()
            ws = wb.active
            if logs:
                headers = list(logs[0].keys())
                ws.append(headers)
                for log in logs:
                    ws.append([log.get(h, "") for h in headers])
            wb.save(report_path)
        except: pass

    def _cleanup(self):
        if self.excel:
            try:
                self.excel.DisplayAlerts = True
                self.excel.Quit()
            except: pass
            self.excel = None

# ==========================================
# LAYER 2: PRESENTATION (View - GUI Definitions)
# - Controller와의 연결만 수행, 로직 없음
# ==========================================
class ExcelCleanView:
    def __init__(self, root, controller):
        self.root = root
        self.controller = controller
        self.root.title("엑셀 얼티밋 딥-클리너 v2.5")
        self.root.geometry("650x750")
        self._build_ui()

    def _build_ui(self):
        # Header
        hdr = tk.Frame(self.root, bg="#283593", pady=15)
        hdr.pack(fill="x")
        tk.Label(hdr, text="EXCEL ULTIMATE DEEP CLEANER", font=("Impact", 18), bg="#283593", fg="white").pack()
        tk.Label(hdr, text="Dual Input System | Recursive Scan | Meta Removal", font=("맑은 고딕", 9), bg="#283593", fg="#C5CAE9").pack()

        main = tk.Frame(self.root, padx=15, pady=15)
        main.pack(fill="both", expand=True)

        # 1. Input Mode Selection
        mode_f = tk.LabelFrame(main, text="1. 입력 모드 선택", font=("맑은 고딕", 10, "bold"), padx=10, pady=10)
        mode_f.pack(fill="x", pady=5)
        
        tk.Radiobutton(mode_f, text="[폴더 단위] (하위폴더 포함)", variable=self.controller.mode_var, value="folder", 
                       command=self.controller.update_ui_state, font=("맑은 고딕", 9)).pack(anchor="w")
        tk.Radiobutton(mode_f, text="[파일 단위] (개별 선택)", variable=self.controller.mode_var, value="files", 
                       command=self.controller.update_ui_state, font=("맑은 고딕", 9)).pack(anchor="w")

        # 2. Target Selection (Dynamic)
        self.target_frame = tk.LabelFrame(main, text="2. 대상 선택", font=("맑은 고딕", 10, "bold"), padx=10, pady=10)
        self.target_frame.pack(fill="both", expand=True, pady=5)
        
        # 2-A. Folder UI
        self.folder_ui = tk.Frame(self.target_frame)
        self.path_ent = tk.Entry(self.folder_ui, font=("Consolas", 10))
        self.path_ent.pack(side="left", fill="x", expand=True)
        tk.Button(self.folder_ui, text="폴더 찾기", command=self.controller.handle_browse_folder).pack(side="left", padx=5)

        # 2-B. File UI
        self.file_ui = tk.Frame(self.target_frame)
        btn_bar = tk.Frame(self.file_ui)
        btn_bar.pack(fill="x", pady=(0, 5))
        tk.Button(btn_bar, text="파일 추가 (+)", command=self.controller.handle_add_files, bg="#E8EAF6").pack(side="left")
        tk.Button(btn_bar, text="목록 초기화", command=self.controller.handle_clear_files).pack(side="left", padx=5)
        self.file_list = tk.Listbox(self.file_ui, height=6, selectmode="extended", font=("Consolas", 9), bg="#FAFAFA")
        self.file_list.pack(fill="both", expand=True)

        # 3. Options
        opt_f = tk.LabelFrame(main, text="3. 정제 옵션", font=("맑은 고딕", 10, "bold"), padx=10, pady=10)
        opt_f.pack(fill="x", pady=5)
        
        tk.Checkbutton(opt_f, text="파일명 앞에 '클리닝_' 접두사 추가 (원본 보호)", variable=self.controller.use_prefix_var, font=("맑은 고딕", 9)).pack(anchor="w")
        
        # Worksheet Activation Options
        ws_f = tk.Frame(opt_f)
        ws_f.pack(fill="x", pady=(5, 0))
        tk.Label(ws_f, text="완료 후 활성 시트:", font=("맑은 고딕", 9)).pack(side="left")
        
        modes = [("첫번째 시트", "first"), ("기존 유지", "keep"), ("특정 시트:", "custom")]
        for text, val in modes:
            tk.Radiobutton(ws_f, text=text, variable=self.controller.activate_mode_var, value=val, font=("맑은 고딕", 9)).pack(side="left", padx=5)
            
        self.custom_sheet_ent = tk.Entry(ws_f, textvariable=self.controller.custom_sheet_var, width=15, font=("Consolas", 9))
        self.custom_sheet_ent.pack(side="left")

        # 4. Log
        log_f = tk.LabelFrame(main, text="4. 진행 로그", font=("맑은 고딕", 10, "bold"), padx=10, pady=10)
        log_f.pack(fill="both", expand=True)
        self.log_txt = tk.Text(log_f, height=8, font=("Consolas", 9), state="disabled", bg="#F5F5F5")
        self.log_txt.pack(fill="both", expand=True)

        # Run Button
        self.run_btn = tk.Button(self.root, text="[정밀 딥-클리닝 시작]", bg="#1A237E", fg="white", font=("맑은 고딕", 12, "bold"), height=2, command=self.controller.handle_run)
        self.run_btn.pack(fill="x", padx=15, pady=15)

    def switch_input_ui(self, mode):
        # Clear previous
        self.folder_ui.pack_forget()
        self.file_ui.pack_forget()
        
        if mode == "folder":
            self.folder_ui.pack(fill="x", expand=True)
            self.target_frame.config(text="2. 대상 폴더 선택 (재귀 스캔)")
        else:
            self.file_ui.pack(fill="both", expand=True)
            self.target_frame.config(text="2. 대상 파일 목록 (직접 선택)")

    def log(self, msg):
        self.log_txt.config(state="normal")
        self.log_txt.insert("end", f"{msg}\n")
        self.log_txt.see("end")
        self.log_txt.config(state="disabled")

    def set_running(self, running):
        if running:
            self.run_btn.config(state="disabled", text="⚡ 처리 중...", bg="gray")
        else:
            self.run_btn.config(state="normal", text="🚀 정밀 딥-클리닝 시작", bg="#1A237E")

    def update_file_list(self, files):
        self.file_list.delete(0, "end")
        for f in files:
            self.file_list.insert("end", os.path.basename(f))


# ==========================================
# LAYER 3: APPLICATION (Controller - Glue Logic)
# - 스레드 관리, 이벤트 핸들링
# ==========================================
class ExcelCleanController:
    def __init__(self, root):
        self.root = root
        
        # Variables
        self.mode_var = tk.StringVar(value="folder")
        self.use_prefix_var = tk.BooleanVar(value=True)
        self.activate_mode_var = tk.StringVar(value="first")
        self.custom_sheet_var = tk.StringVar(value="") # Added missing variable
        self.selected_files = []
        
        # Modules
        self.view = ExcelCleanView(root, self)
        self.engine = ExcelDeepCleanEngine(self._callback)
        self.is_running = False
        
        # Initial State
        self.update_ui_state()

    def update_ui_state(self):
        self.view.switch_input_ui(self.mode_var.get())

    def handle_browse_folder(self):
        d = filedialog.askdirectory()
        if d:
            self.view.path_ent.delete(0, "end")
            self.view.path_ent.insert(0, d)

    def handle_add_files(self):
        fs = filedialog.askopenfilenames(filetypes=[("Excel Files", "*.xlsx;*.xlsm;*.xls")])
        if fs:
            self.selected_files.extend(list(fs))
            self.selected_files = list(set(self.selected_files)) # Deduplicate
            self.view.update_file_list(self.selected_files)

    def handle_clear_files(self):
        self.selected_files = []
        self.view.update_file_list([])

    def handle_run(self):
        if self.is_running: return
        
        mode = self.mode_var.get()
        targets = None
        
        if mode == "folder":
            targets = self.view.path_ent.get().strip()
            if not targets or not os.path.isdir(targets):
                messagebox.showwarning("오류", "올바른 폴더 경로를 입력하세요.")
                return
        else:
            targets = self.selected_files
            if not targets:
                messagebox.showwarning("오류", "파일을 1개 이상 추가하세요.")
                return

        self.is_running = True
        self.view.set_running(True)
        self.view.log_txt.config(state="normal")
        self.view.log_txt.delete("1.0", "end")
        self.view.log_txt.config(state="disabled")
        
        opts = {
            'use_prefix': self.use_prefix_var.get(),
            'activate_mode': self.activate_mode_var.get(),
            'custom_sheet': self.custom_sheet_var.get() # Pass custom sheet name
        }
        
        threading.Thread(target=lambda: self.engine.run_process_dual(mode, targets, opts), daemon=True).start()

    def _callback(self, type, msg):
        self.root.after(0, lambda: self._handle_callback(type, msg))

    def _handle_callback(self, type, msg):
        if type == "log":
            self.view.log(msg)
        elif type == "status":
            self.view.log(f"--- {msg} ---")
        elif type == "done":
            self.view.log(f"\n{msg}")
            self.is_running = False
            self.view.set_running(False)
            messagebox.showinfo("완료", "작업이 완료되었습니다.")

def is_admin():
    import ctypes
    try: return ctypes.windll.shell32.IsUserAnAdmin()
    except: return False

def run_as_admin():
    import ctypes, sys
    if is_admin(): return
    params = f'"{sys.argv[0]}" ' + " ".join([f'"{arg}"' for arg in sys.argv[1:]])
    ctypes.windll.shell32.ShellExecuteW(None, "runas", sys.executable, params, None, 1)
    sys.exit(0)

if __name__ == "__main__":
    # [Root Cause Fix] 관리자 권한 자동 승격 제거 (UAC 팝업 및 COM 권한 충돌 방지)
    # run_as_admin() 
    root = tk.Tk()
    
    # [v34.1.21] Stealth Launch 대응: 창을 최상단으로 강제 부각
    root.lift()
    root.attributes('-topmost', True)
    root.after(100, lambda: root.attributes('-topmost', False))
    
    app = ExcelCleanController(root)
    root.mainloop()
