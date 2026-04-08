"""
================================================================================
 [EXCEL] 엑셀 금액 무결성 심사 도구 (Excel Amount Integrity Checker) v34.1.16
================================================================================
 - Clean Layer Architecture: Engine(Domain) / View(Presentation) / Controller(Application)
 - 특정 셀(M32) 기반 금액 검증 및 접두사 부여
 - PRD 준수: 단일 파일 유지, 기존 로직 100% 보존
 - 무결성 보증: sys.stdout UTF-8 강제 설정 및 이모지 제거 (CP949 호환)
================================================================================
"""
import sys
try:
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8')
except: pass
import os
import openpyxl
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from pathlib import Path
import threading


# ==========================================
# 1. DOMAIN LAYER (Core Engine - 비즈니스 로직)
# ==========================================
class AmountCheckEngine:
    """금액 무결성 점검 핵심 엔진: GUI와 분리된 순수 로직"""
    
    @staticmethod
    def check_cell_validity(file_path, sheet_name, cell_addr):
        """
        특정 셀의 값이 유효한 금액(0이 아닌 숫자)인지 확인
        Returns: (is_valid, value)
        """
        try:
            wb = openpyxl.load_workbook(str(file_path), data_only=True, read_only=True)
            if sheet_name not in wb.sheetnames:
                wb.close()
                return None, "시트 없음"
            
            ws = wb[sheet_name]
            try:
                val = ws[cell_addr].value
            except:
                val = None
            wb.close()
            
            # 핵심 로직: 유효한 금액은 0이 아닌 숫자여야 함
            is_valid = isinstance(val, (int, float)) and val != 0
            return is_valid, val
        except Exception as e:
            return None, str(e)
    
    @staticmethod
    def process_files(files, sheet_name, cell_addr, prefix, callback=None):
        """
        파일들의 금액 무결성 점검 및 접두사 부여 수행
        callback: (type, message) 형태로 진행 상황 전달
        """
        processed = 0
        renamed = 0
        errors = []
        skipped = 0
        
        for f_path in files:
            file_path = Path(f_path)
            processed += 1
            filename = file_path.name
            
            if callback:
                callback("progress", f"검사 중: {filename}")
            
            is_valid, val = AmountCheckEngine.check_cell_validity(file_path, sheet_name, cell_addr)
            
            if is_valid is None:
                errors.append(f"{filename}: {val}")
                if callback:
                    callback("error", f"  [오류] {filename}: {val}")
                continue
            
            if not is_valid:
                # 부실 파일 처리
                if not filename.startswith(prefix):
                    new_name = prefix + filename
                    new_path = file_path.with_name(new_name)
                    
                    if not new_path.exists():
                        try:
                            os.rename(str(file_path), str(new_path))
                            renamed += 1
                            if callback:
                                callback("renamed", f"  [부실→변경] {filename} → {new_name} (값: {val})")
                        except Exception as e:
                            errors.append(f"{filename}: {str(e)}")
                            if callback:
                                callback("error", f"  [변경 실패] {filename}: {str(e)}")
                    else:
                        skipped += 1
                        if callback:
                            callback("skip", f"  [스킵] 이미 존재: {new_name}")
                else:
                    skipped += 1
                    if callback:
                        callback("skip", f"  [스킵] 이미 접두사 있음: {filename}")
            else:
                if callback:
                    callback("valid", f"  [정상] {filename} (값: {val})")
        
        return {
            "processed": processed,
            "renamed": renamed,
            "skipped": skipped,
            "errors": errors
        }


# ==========================================
# 2. PRESENTATION LAYER (View - GUI 정의)
# ==========================================
class AmountCheckView:
    """GUI 화면 구성: Controller와 분리된 순수 View"""
    
    def __init__(self, root, controller):
        self.root = root
        self.controller = controller
        self.root.title("[엑셀 금액 무결성 심사 도구 v34.1.16]")
        self.root.geometry("700x650")
        self._build_ui()
    
    def _build_ui(self):
        # 헤더
        header = tk.Frame(self.root, bg="#D32F2F", pady=12)
        header.pack(fill="x")
        tk.Label(header, text="[엑셀 금액 무결성 심사 도구 v34.1.16]", font=("Malgun Gothic", 14, "bold"),
                 bg="#D32F2F", fg="white").pack()
        tk.Label(header, text="특정 셀 기반 금액 검증 및 부실 파일 접두사 부여 (Clean Architecture)",
                 font=("Malgun Gothic", 9), bg="#D32F2F", fg="#FFCDD2").pack()
        
        main = tk.Frame(self.root, padx=20, pady=10)
        main.pack(fill="both", expand=True)
        
        # 1. 파일 선택
        box1 = tk.LabelFrame(main, text="[1] 검사 대상 파일 선택", font=("Malgun Gothic", 10, "bold"), padx=10, pady=10)
        box1.pack(fill="x", pady=5)
        btn_f = tk.Frame(box1)
        btn_f.pack(fill="x")
        self.select_btn = tk.Button(btn_f, text="[파일 선택]", command=self.controller.handle_select_files,
                                    font=("Malgun Gothic", 10), bg="#FFEBEE", width=15)
        self.select_btn.pack(side="left")
        self.file_count_label = tk.Label(btn_f, text="선택된 파일: 0개", font=("Malgun Gothic", 9))
        self.file_count_label.pack(side="left", padx=15)
        
        # 2. 검사 조건 설정
        box2 = tk.LabelFrame(main, text="[2] 검사 조건 설정", font=("Malgun Gothic", 10, "bold"), padx=10, pady=10)
        box2.pack(fill="x", pady=5)
        
        f2a = tk.Frame(box2)
        f2a.pack(fill="x", pady=3)
        tk.Label(f2a, text="워크시트 이름:", font=("Malgun Gothic", 9), width=15, anchor="e").pack(side="left")
        tk.Entry(f2a, textvariable=self.controller.sheet_name_var, font=("Malgun Gothic", 10), width=25).pack(side="left", padx=5)
        
        f2b = tk.Frame(box2)
        f2b.pack(fill="x", pady=3)
        tk.Label(f2b, text="검사 셀 주소:", font=("Malgun Gothic", 9), width=15, anchor="e").pack(side="left")
        tk.Entry(f2b, textvariable=self.controller.cell_addr_var, font=("Malgun Gothic", 10), width=10).pack(side="left", padx=5)
        
        f2c = tk.Frame(box2)
        f2c.pack(fill="x", pady=3)
        tk.Label(f2c, text="부실 파일 접두사:", font=("Malgun Gothic", 9), width=15, anchor="e").pack(side="left")
        tk.Entry(f2c, textvariable=self.controller.prefix_var, font=("Malgun Gothic", 10), width=10).pack(side="left", padx=5)
        
        # 판단 기준 안내
        info_f = tk.Frame(box2, bg="#FFF3E0")
        info_f.pack(fill="x", pady=10)
        tk.Label(info_f, text="[INFO] 부실 처리 대상 (접두사가 붙는 경우):", font=("Malgun Gothic", 9, "bold"), bg="#FFF3E0").pack(anchor="w")
        tk.Label(info_f, text="  • 숫자가 아닌 모든 값 (문자열 '-', IFERROR 결과 등)", font=("Malgun Gothic", 8), bg="#FFF3E0").pack(anchor="w")
        tk.Label(info_f, text="  • 숫자 0 (회계 서식 포함)", font=("Malgun Gothic", 8), bg="#FFF3E0").pack(anchor="w")
        tk.Label(info_f, text="  • 데이터 누락 (빈 셀) 및 수식 오류 (#VALUE! 등)", font=("Malgun Gothic", 8), bg="#FFF3E0").pack(anchor="w")
        
        # 3. 진행 로그
        box3 = tk.LabelFrame(main, text="[3] 진행 상황", font=("Malgun Gothic", 10, "bold"), padx=10, pady=10)
        box3.pack(fill="both", expand=True, pady=5)
        
        self.log_txt = tk.Text(box3, font=("Consolas", 9), bg="#FAFAFA", height=10)
        scr = tk.Scrollbar(box3, command=self.log_txt.yview)
        self.log_txt.config(yscrollcommand=scr.set)
        scr.pack(side="right", fill="y")
        self.log_txt.pack(fill="both", expand=True)
        
        # 실행 버튼
        self.run_btn = tk.Button(self.root, text="[심층 금액 점검 실행]", font=("Malgun Gothic", 12, "bold"),
                                 bg="#C62828", fg="white", height=2, command=self.controller.handle_start)
        self.run_btn.pack(fill="x", padx=20, pady=15)
        
        # 상태바
        self.status_bar = tk.Label(self.root, text="준비됨 | 파일을 선택하고 실행 버튼을 누르세요.",
                                   bd=1, relief="sunken", anchor="w", font=("Malgun Gothic", 8), padx=10)
        self.status_bar.pack(side="bottom", fill="x")
    
    def log(self, msg):
        self.log_txt.insert("end", f"{msg}\n")
        self.log_txt.see("end")
    
    def clear_log(self):
        self.log_txt.delete("1.0", "end")
    
    def set_status(self, msg):
        self.status_bar.config(text=msg)
    
    def update_file_count(self, count):
        self.file_count_label.config(text=f"선택된 파일: {count}개")
    
    def set_running_state(self, is_running):
        if is_running:
            self.run_btn.config(state="disabled", bg="#9E9E9E")
            self.select_btn.config(state="disabled")
        else:
            self.run_btn.config(state="normal", bg="#C62828")
            self.select_btn.config(state="normal")


# ==========================================
# 3. APPLICATION LAYER (Controller - 로직 연결)
# ==========================================
class AmountCheckController:
    """Controller: View와 Engine을 연결하고 사용자 이벤트 처리"""
    
    def __init__(self, root):
        self.root = root
        self.engine = AmountCheckEngine()
        self.is_running = False
        self.selected_files = []
        
        # 상태 변수 초기화
        self.sheet_name_var = tk.StringVar(value="PR당사안")
        self.cell_addr_var = tk.StringVar(value="M32")
        self.prefix_var = tk.StringVar(value="F")
        
        # View 생성
        self.view = AmountCheckView(root, self)
    
    def handle_select_files(self):
        """파일 선택 처리"""
        files = filedialog.askopenfilenames(
            title="금액 무결성 점검 대상 파일을 선택하세요",
            filetypes=[("Excel files", "*.xlsx *.xlsm")]
        )
        if files:
            self.selected_files = list(files)
            self.view.update_file_count(len(self.selected_files))
            self.view.log(f"[*] {len(self.selected_files)}개 파일 선택됨")
    
    def handle_start(self):
        """금액 점검 실행"""
        if self.is_running:
            return
        
        if not self.selected_files:
            messagebox.showwarning("경고", "파일이 선택되지 않았습니다.")
            return
        
        sheet_name = self.sheet_name_var.get().strip()
        cell_addr = self.cell_addr_var.get().strip()
        prefix = self.prefix_var.get()
        
        if not sheet_name or not cell_addr:
            messagebox.showwarning("경고", "시트 이름과 셀 주소를 입력해주세요.")
            return
        
        # [Root Cause Fix] 실행 전 확인 다이얼로그 제거 (대시보드 직결 실행 대응)
        # confirm = messagebox.askyesno(
        #     "확인",
        #     f"심층 금액 점검을 시작합니다.\n\n"
        #     f"대상: {len(self.selected_files)}개 파일\n"
        #     f"위치: {sheet_name} > {cell_addr}\n\n"
        #     f"부실 파일에 '{prefix}' 접두사가 붙습니다.\n"
        #     f"진행하시겠습니까?"
        # )
        # if not confirm:
        #     return
        
        self.is_running = True
        self.view.set_running_state(True)
        self.view.clear_log()
        self.view.log(f"[INFO] 금액 점검 시작 (대상: {len(self.selected_files)}개)")
        self.view.log("-" * 50)
        self.view.set_status("검사 중...")
        
        def run():
            result = self.engine.process_files(
                self.selected_files,
                sheet_name,
                cell_addr,
                prefix,
                self._callback
            )
            self.root.after(0, lambda: self._finalize(result))
        
        threading.Thread(target=run, daemon=True).start()
    
    def _callback(self, msg_type, msg):
        self.root.after(0, lambda: self.view.log(msg))
    
    def _finalize(self, result):
        self.is_running = False
        self.view.set_running_state(False)
        
        self.view.log("-" * 50)
        self.view.log(f"[OK] 점검 완료!")
        self.view.log(f"   총 검사: {result['processed']}개")
        self.view.log(f"   부실→이름변경: {result['renamed']}개")
        self.view.log(f"   스킵: {result['skipped']}개")
        self.view.log(f"   오류: {len(result['errors'])}개")
        self.view.set_status(f"완료: 변경 {result['renamed']}개, 오류 {len(result['errors'])}개")
        
        report = f"심층 금액 점검 완료!\n\n총 검사: {result['processed']}\n부실/비수량 파일 수정: {result['renamed']}"
        if result['errors']:
            report += f"\n\n오류: {len(result['errors'])}건"
        messagebox.showinfo("최종 리포트", report)


# ==========================================
# 4. ENTRY POINT
# ==========================================
if __name__ == "__main__":
    root = tk.Tk()

    # [v34.1.21] Stealth Launch 대응: 창을 최상단으로 강제 부각
    root.lift()
    root.attributes('-topmost', True)
    root.after(100, lambda: root.attributes('-topmost', False))

    AmountCheckController(root)
    root.mainloop()
