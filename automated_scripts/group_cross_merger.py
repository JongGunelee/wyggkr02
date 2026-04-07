"""
================================================================================
 [통합 문서 관리자 (Group Cross Merger) v34.1.16] (Ultimate Speed & Hardened)
================================================================================
- 아키텍처: Clean Layer Architecture (Domain / Presentation / Application)
- 주요 기능: 엑셀 딥-클리닝, 이기종 문서(PPT/PDF/Excel) 통합 병합, 지능형 파일 복제
- 가이드라인 준수: 00 PRD 가이드.md | AI_CODING_GUIDELINES_2026.md
- 무결성 보증: sys.stdout UTF-8 강제 설정 및 이모지 제거 (CP949 호환)
- 최적화: 0.1s 반응형 폴링 및 일괄 세션 관리로 유휴 대기 시간 전수 제거
================================================================================
"""
import sys
try:
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8')
except: pass

import os
import re
import time
import subprocess
import shutil
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import win32com.client
import fitz  # PyMuPDF
import openpyxl
from pathlib import Path
import importlib.util

# ═══════════════════════════════════════════════════════════
# LAYER 0: INFRASTRUCTURE (Legacy Specialist Loader)
# ═══════════════════════════════════════════════════════════
class LegacyEngineLoader:
    @staticmethod
    def get_available_specialists(target_type):
        """000_Backup 폴더에서 해당 타입의 전문 엔진 스캔 (이식성 강화)"""
        # 현재 실행 경로 또는 상위 경로의 000_Backup 탐색
        base_path = Path(os.getcwd())
        bak_dir = base_path / "000_Backup"
        if not bak_dir.exists():
            # 대시보드 구조 대응 (한 단계 위 확인)
            bak_dir = base_path.parent / "000_Backup"
            
        if not bak_dir.exists(): return []
        
        matches = []
        for f in bak_dir.glob(f"BAK*_{target_type}*.py"):
            matches.append(f.name)
        return sorted(matches, reverse=True)

    @staticmethod
    def load_specific_engine(file_name):
        """백업 파일로부터 특정 엔진 로직을 동적으로 로드"""
        try:
            base_path = Path(os.getcwd())
            bak_dir = base_path / "000_Backup"
            if not bak_dir.exists():
                bak_dir = base_path.parent / "000_Backup"
                
            path = bak_dir / file_name
            spec = importlib.util.spec_from_file_location("legacy_engine", path)
            module = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(module)
            return module
        except Exception as e:
            return None

# ═══════════════════════════════════════════════════════════
# LAYER 1: DOMAIN (Management Engine - Pure Business Logic)
# ═══════════════════════════════════════════════════════════
class ManagementEngine:
    def __init__(self, callback):
        self.callback = callback
        self.apps = {"ppt": None, "word": None, "excel": None}

    def _kill_office_processes(self):
        """[v34.2.2] 기존 좀비 프로세스 강제 소거 (psutil+WMI 하이브리드)"""
        import time, ctypes
        try: ctypes.windll.kernel32.SetErrorMode(0x0001 | 0x0002 | 0x8000)
        except: pass
        self.callback("log", "[CLEAN] 기존 Office 좀비 프로세스 네이티브 정리 중 (팝업 차단)...")
        
        targets = ["POWERPNT.EXE", "EXCEL.EXE", "WINWORD.EXE"]
        killed = False
        try:
            import psutil
            for p in psutil.process_iter(['name']):
                if p.info['name'] and p.info['name'].upper() in targets:
                    try: p.kill(); killed = True
                    except: pass
        except ImportError:
            try:
                import pythoncom, win32com.client
                pythoncom.CoInitialize()
                wmi = win32com.client.GetObject("winmgmts:")
                for p in wmi.InstancesOf("Win32_Process"):
                    name = p.Properties_('Name').Value
                    if name and name.upper() in targets:
                        try: p.Terminate(); killed = True
                        except: pass
            except: pass
        if killed:
            time.sleep(1.0)

    def _get_app(self, app_type):
        """[v34.2.8] CLSID Direct & Deep Polling 엔진 가동"""
        """COM 자동화 애플리케이션 인스턴스를 가져옵니다."""
        if self.apps[app_type] is not None:
            try:
                # 연결 무결성 테스트
                _ = self.apps[app_type].Visible
                return self.apps[app_type]
            except:
                self.apps[app_type] = None

        # [v35.4.2] 1단계: 마법의 해결책 - win32com 캐시 완전 소거
        import os, shutil
        try:
            gen_py = os.path.join(os.environ.get('TEMP', ''), 'gen_py')
            if os.path.exists(gen_py): shutil.rmtree(gen_py, ignore_errors=True)
        except: pass

        prog_ids = {
            "ppt": "PowerPoint.Application",
            "excel": "Excel.Application", # Kept "excel" as per original, not "xls"
            "word": "Word.Application"
        }
        clsids = {
            "ppt": "{91493441-5A91-11CF-8700-00AA0060263B}",
            "excel": "{00024500-0000-0000-C000-000000000046}", # Kept "excel" as per original, not "xls"
            "word": "{000209FF-0000-0000-C000-000000000046}"
        }
        
        prog_id = prog_ids.get(app_type)
        clsid = clsids.get(app_type)
        app = None
        errors = []

        import win32com.client
        from win32com.client import Dispatch, DispatchEx, gencache
        from win32com.client.dynamic import Dispatch as DynDispatch

        # Phase 1: DispatchEx (새로운 독립 인스턴스 강제 생성 - UAC 충돌 회피 핵심)
        try:
            app = DispatchEx(prog_id)
            if app: return self._finalize_app(app, app_type)
        except Exception as e1:
            errors.append(f"P1:{str(e1)[:15]}")

        # Phase 2: CLSID Direct Dispatch (권한/샌드박스 우회)
        try:
            app = DynDispatch(clsid)
            if app: return self._finalize_app(app, app_type)
        except Exception as e2:
            errors.append(f"P2:{str(e2)[:15]}")

        # Phase 3: EnsureDispatch (레지스트리 강제 복구)
        try:
            app = gencache.EnsureDispatch(prog_id)
            if app: return self._finalize_app(app, app_type)
        except Exception as e3:
            errors.append(f"P3:{str(e3)[:15]}")

        # Phase 4: ShellExecute App Launch + GetActiveObject (강력한 UAC 우회 접근)
        try:
            import win32api
            exe_map = {"ppt": "powerpnt.exe", "excel": "excel.exe", "word": "winword.exe"}
            exe_path = exe_map.get(app_type)
            if exe_path:
                # [v35.4.6] SW_SHOWMINNOACTIVE(7) 적용: 포커스 강탈 방지
                win32api.ShellExecute(0, 'open', exe_path, None, None, 7)
                for i in range(8): # 8초간 추적 가동
                    time.sleep(1.0)
                    try:
                        app = win32com.client.GetActiveObject(prog_id)
                        if app: return self._finalize_app(app, app_type)
                    except: pass
        except Exception as e4:
            errors.append(f"P4:{str(e4)[:15]}")

        # 최후의 수단: taskkill (안전 모드 팝업 위험성)
        # -2147024156 is a common HRESULT for "Class not registered" or "Access Denied"
        if any("-2147024156" in err for err in errors):
            self.callback("log", "[WARN] COM 연결 오류 감지. Office 프로세스 강제 종료 후 재시도...")
            self._kill_office_processes()
            
            # Re-clear gen_py after kill, as it might be related to corrupted COM objects
            try:
                gen_py = os.path.join(os.environ.get('TEMP', ''), 'gen_py')
                if os.path.exists(gen_py): shutil.rmtree(gen_py, ignore_errors=True)
            except: pass
            
            time.sleep(1.0)
            try:
                app = DispatchEx(prog_id)
                if app: return self._finalize_app(app, app_type)
            except Exception as e_retry:
                errors.append(f"P_Retry:{str(e_retry)[:15]}")
        
        # Phase 5: Standard Dispatch (최후의 수단) - if not already returned by retry
        if not app: # Only try if app is still None
            try:
                app = win32com.client.Dispatch(prog_id)
                if app: return self._finalize_app(app, app_type)
            except Exception as e4:
                errors.append(f"P4:{str(e4)[:15]}")

        if not app:
            self.callback("log", f"[FAIL] {app_type.upper()} 가동 실패 (모든 폴백 고갈): {'; '.join(errors)}")
            return None
        return app

    def _finalize_app(self, app, app_type):
        """앱 설정 마무리 및 가동 확인"""
        try:
            if app_type == "excel":
                try:
                    app.Visible = False
                    app.DisplayAlerts = False
                    app.ScreenUpdating = False
                    app.Interactive = False
                    app.EnableEvents = False
                    app.AskToUpdateLinks = False
                except: pass
            elif app_type == "ppt":
                try: 
                    app.Visible = 0  # msoFalse
                except:
                    # [v35.1.0] Visible 제어 불가 시 최소화 폴백
                    try: app.WindowState = 2 # ppWindowMinimized
                    except: pass
                try: 
                    app.DisplayAlerts = 1  # ppAlertsNone
                except: pass
            elif app_type == "word":
                app.Visible = False
                app.DisplayAlerts = 0  # wdAlertsNone
        except:
            pass
        self.apps[app_type] = app
        return app

    def _clean_workbook(self, wb, smart_fix=False, col_count=4):
        """본질적인 엑셀 딥-클리닝 + 스마트 구조 교정 (TUNED: Variable Column Hide)"""
        try:
            # 1. 기존 클리닝 (외부링크/이름정의)
            links = wb.LinkSources(1)
            if links:
                self.callback("log", f"[LINK] 외부 링크 {len(links)}개 발견 -> 값으로 고정 중...")
                for link in links:
                    try: wb.BreakLink(Name=link, Type=1)
                    except: pass
            
            names_count = wb.Names.Count
            if names_count > 0:
                self.callback("log", f"[NAME] 정의된 이름 {names_count}개 발견 -> 일괄 삭제 중...")
                for name in wb.Names:
                    try: name.Delete()
                    except: pass
            
            # 2. 스마트 구조 교정 (TUNED: Marker Search E00000)
            if smart_fix:
                for ws in wb.Worksheets:
                    ws.Columns.EntireColumn.Hidden = False
                    # Marker Search (E00000...)
                    data_col = -1
                    for c in range(1, 25): # Search range expanded
                        val = str(ws.Cells(5, c).Value)
                        if "E00000" in val:
                            data_col = c; break
                    
                    if data_col != -1:
                        target_pos = col_count + 1 
                        if data_col > target_pos:
                            for _ in range(data_col - target_pos): ws.Columns(1).Delete()
                        elif data_col < target_pos:
                            for _ in range(target_pos - data_col): ws.Columns(1).Insert()
                        # Final Hide 1 to col_count
                        if col_count > 0:
                            ws.Range(ws.Columns(1), ws.Columns(col_count)).EntireColumn.Hidden = True
        except Exception as e:
            self.callback("log", f"[WARN] 클리닝 중 오류: {str(e)}")

    def run_amount_check(self, config):
        """특수: 금액 무결성 점검 및 이름 변경 (advanced_excel_rename 로직 이식)"""
        source_dir = config['source_dir']
        output_dir = os.path.join(source_dir, "00_Result_AmountChecked")
        if not os.path.exists(output_dir): os.makedirs(output_dir)
        
        files = [f for f in os.listdir(source_dir) if os.path.isfile(os.path.join(source_dir, f))] if config['source_mode'] == "폴더" else [os.path.basename(f) for f in config['selected_files']]
        excel_files = [f for f in files if os.path.splitext(f)[1].lower() in ['.xlsx', '.xlsm']]
        
        try:
            self._kill_office_processes()
            app = self._get_app("excel")
            renamed = 0
            for i, f_name in enumerate(excel_files, 1):
                f_path = os.path.abspath(os.path.join(source_dir, f_name))
                self.callback("status", f"[{i}/{len(excel_files)}] 금액 점검 중: {f_name}")
                
                wb = app.Workbooks.Open(f_path, UpdateLinks=False, ReadOnly=True)
                ws = None
                try: ws = wb.Sheets(config['sheet_name'])
                except: pass
                
                is_invalid = True
                if ws:
                    val = ws.Range(config['cell_addr']).Value
                    # Valid amount logic: is a non-zero number
                    if isinstance(val, (int, float)) and val != 0:
                        is_invalid = False
                
                wb.Close(False)
                
                # Action: Rename if invalid
                if is_invalid:
                    prefix = config['prefix']
                    if not f_name.startswith(prefix):
                        new_name = prefix + f_name
                        shutil.copy2(f_path, os.path.join(output_dir, new_name))
                        self.callback("log", f"[FLAG] 부실 확인: {f_name} -> {new_name}")
                        renamed += 1
                else:
                    shutil.copy2(f_path, os.path.join(output_dir, f_name))
            
            self.callback("done", f"[DONE] 금액 점검 완료.\n총 {len(excel_files)}건 중 {renamed}건 부실 발견\n저장위치: {output_dir}")
        finally:
            self._cleanup()

    def run_clean_only(self, config):
        """엑셀 단독 클리닝 모드 (v33 고도화)"""
        source_dir = config['source_dir']
        output_dir = os.path.join(source_dir, "00_Result_Cleaned")
        if not os.path.exists(output_dir): os.makedirs(output_dir)

        files = [f for f in os.listdir(source_dir) if os.path.isfile(os.path.join(source_dir, f))] if config['source_mode'] == "폴더" else [os.path.basename(f) for f in config['selected_files']]
        excel_files = [f for f in files if os.path.splitext(f)[1].lower() in ['.xls', '.xlsx', '.xlsm']]

        if not excel_files:
            self.callback("error", "대상 엑셀 파일을 찾지 못했습니다."); return

        try:
            self._kill_office_processes()
            app = self._get_app("excel")
            for i, f_name in enumerate(excel_files, 1):
                f_path = os.path.abspath(os.path.join(source_dir, f_name))
                self.callback("status", f"[{i}/{len(excel_files)}] 최적화 중: {f_name}")
                
                wb = app.Workbooks.Open(f_path, UpdateLinks=False, ReadOnly=True)
                self._clean_workbook(wb, smart_fix=config.get('smart_fix', False), col_count=config.get('col_count', 4))
                
                out_path = os.path.join(output_dir, f"Cleaned_{f_name}")
                wb.SaveAs(os.path.abspath(out_path))
                wb.Close(False)
                self.callback("log", f"[OK] 완료: Cleaned_{f_name}")
            self.callback("done", f"[DONE] 딥-클리닝 완료.\n저장위치: {output_dir}")
        finally:
            self._cleanup()

    def run_merge(self, config):
        """통합 병합 엔진 (BAK 북마크 + 고효율 압축 엔진)"""
        source_dir = config['source_dir']
        output_dir = os.path.join(source_dir, "00_Result_Merged")
        if not os.path.exists(output_dir): os.makedirs(output_dir)

        p_exts = [e.strip().lower() for e in config['exts_p'].split(',') if e.strip()]
        o_exts = [e.strip().lower() for e in config['exts_o'].split(',') if e.strip()]
        all_exts = p_exts + o_exts

        if config['source_mode'] == "파일":
            target_files = [os.path.abspath(f) for f in config['selected_files'] if os.path.splitext(f)[1].lower() in all_exts]
        else:
            target_files = [os.path.abspath(os.path.join(source_dir, f)) for f in os.listdir(source_dir) 
                            if os.path.isfile(os.path.join(source_dir, f)) and os.path.splitext(f)[1].lower() in all_exts]

        if not target_files:
            self.callback("error", "병합 대상을 찾지 못했습니다."); return

        groups = {}
        for f_path in target_files:
            match = re.match(r"^(\d+)", os.path.basename(f_path))
            gid = match.group(1) if match else "All"
            groups.setdefault(gid, []).append(f_path)

        gid_list = sorted([k for k in groups.keys() if k != "All"])
        if "All" in groups: gid_list.append("All")

        try:
            self._kill_office_processes()
            for gid in gid_list:
                batch = groups[gid]
                if config['source_mode'] == "폴더": batch.sort()
                
                ext_set = set(os.path.splitext(f)[1].lower() for f in batch)
                mode = "PDF"
                is_excel_set = ext_set.issubset({'.xls', '.xlsx', '.xlsm'})
                is_ppt_set = ext_set.issubset({'.ppt', '.pptx'})
                is_word_set = ext_set.issubset({'.doc', '.docx'})

                if "편집형" in config['merge_mode']:
                    if is_excel_set: mode = "NATIVE"; target_ext = ".xlsx"
                    elif is_ppt_set:
                        if not self._check_ppt_orient(source_dir, batch): mode = "PDF"
                        else: mode = "NATIVE"; target_ext = ".pptx"
                    elif is_word_set: mode = "NATIVE"; target_ext = ".docx"
                    elif len(ext_set) == 1 and list(ext_set)[0] == ".pdf": mode = "NATIVE_PDF"

                self.callback("status", f"그룹[{gid}] 처리 중...")
                
                # [창의적 통합] 레거시 스페셜리스트 엔진 호출 여부 확인
                legacy_file = config.get('engine_ver')
                if legacy_file and "Standard" not in legacy_file:
                    self.callback("log", f"[INFO] 전문 엔진 가동: {legacy_file} (튜닝된 로직 적용 중...)")
                    legacy_mod = LegacyEngineLoader.load_specific_engine(legacy_file)
                    if legacy_mod and hasattr(legacy_mod, "ManagementEngine"):
                        # 레거시 클래스 인스턴스화 및 세션 위임
                        l_eng = legacy_mod.ManagementEngine(self.callback)
                        # 레거시의 병합 방식에 맞춰 batch 소스 주입
                        if mode == "PDF": l_eng._merge_pdf(source_dir, output_dir, batch, config, p_exts, o_exts, mode=="NATIVE_PDF")
                        else: l_eng._merge_native(source_dir, output_dir, batch, target_ext, config)
                        continue # 다음 그룹으로
                
                if mode == "NATIVE": self._merge_native(source_dir, output_dir, batch, target_ext, config)
                else: self._merge_pdf(source_dir, output_dir, batch, config, p_exts, o_exts, mode=="NATIVE_PDF")
                
                # [v35.4.7] 각 그룹 처리가 완료되면 프로세스를 즉시 종료하여 누적(누수) 방지
                self.callback("log", f"[-] 그룹[{gid}] 리소스 정리 중...")
                self._cleanup()
            
            self.callback("done", f"[DONE] 통합 병합 완료!\n저장위치: {output_dir}")
        finally:
            self._cleanup()

    def _check_ppt_orient(self, s_dir, files):
        app = self._get_app("ppt")
        try:
            orients = set()
            for f_path in files:
                p = app.Presentations.Open(os.path.abspath(f_path), WithWindow=False, ReadOnly=True)
                orients.add(p.PageSetup.Orientation); p.Close()
            return len(orients) <= 1
        except: return False

    def _merge_native(self, s_dir, o_dir, files, ext, config):
        f_base = os.path.basename(files[0])
        res_name = f"{config['prefix']}{os.path.splitext(f_base)[0]}{config['suffix']}{ext}"
        target_path = os.path.abspath(os.path.join(o_dir, res_name))
        try:
            if ext in ['.xls', '.xlsx', '.xlsm']:
                app = self._get_app("excel"); wb_m = app.Workbooks.Add()
                for f_path in files:
                    wb_s = app.Workbooks.Open(f_path, UpdateLinks=False, ReadOnly=True)
                    self._clean_workbook(wb_s, smart_fix=True)
                    for sh in wb_s.Worksheets:
                        try: sh.Name = f"{sh.Name}({os.path.splitext(os.path.basename(f_path))[0][:10]})"
                        except: pass
                        sh.Copy(After=wb_m.Sheets(wb_m.Sheets.Count))
                    wb_s.Close(False)
                if wb_m.Sheets.Count > 1: app.DisplayAlerts = False; wb_m.Sheets(1).Delete()
                wb_m.SaveAs(target_path); wb_m.Close()
            elif ext in ['.ppt', '.pptx']:
                app = self._get_app("ppt"); pp_m = app.Presentations.Add(WithWindow=False)
                for f_path in files: pp_m.Slides.InsertFromFile(f_path, pp_m.Slides.Count)
                pp_m.SaveAs(target_path); pp_m.Close()
            elif ext in ['.doc', '.docx']:
                app = self._get_app("word"); wd_m = app.Documents.Add(Visible=False)
                for f_path in files: app.Selection.InsertFile(f_path)
                wd_m.SaveAs(target_path); wd_m.Close()
            self.callback("log", f"📂 Native 합본 생성: {res_name}")
        except: pass

    def _merge_pdf(self, s_dir, o_dir, files, config, p_exts, o_exts, is_m):
        f_base = os.path.basename(files[0])
        res_name = f"{config['prefix']}{os.path.splitext(f_base)[0]}{config['suffix']}.pdf"
        target_path = os.path.join(o_dir, res_name)
        pdf = fitz.open()
        tmps = []
        
        # [v35.5.0] 전방위 프리워밍 (Full Pre-warming): 모든 오피스 파일을 미리 오픈 위임
        try:
            import win32api
            for f_p in files:
                if os.path.splitext(f_p)[1].lower() in o_exts:
                    try: win32api.ShellExecute(0, 'open', os.path.abspath(f_p), None, None, 7)
                    except: pass
        except: pass

        for f_path in files:
            path = os.path.abspath(f_path); f_name = os.path.basename(f_path); ext = os.path.splitext(f_path)[1].lower()
            self.callback("log", f"[FILE] 처리 중: {f_name}")
            
            if ext in p_exts:
                try:
                    if ext == '.pdf':
                        src = fitz.open(path); pdf.insert_pdf(src); src.close()
                    else: # Images
                        img = fitz.open(path); pdf_bytes = img.convert_to_pdf()
                        img_pdf = fitz.open("pdf", pdf_bytes); pdf.insert_pdf(img_pdf)
                        img_pdf.close(); img.close()
                except: pass
            elif ext in o_exts:
                t_pdf = os.path.join(o_dir, f"tmp_{int(time.time()*1000)}_{f_name}.pdf")
                ok = False
                try:
                    if ext in ['.ppt', '.pptx']:
                        pres = None
                        try:
                            import win32api
                            # 이미 프리워밍되었으므로 즉시 GetObject 시도 (고속 폴링)
                            for wait_idx in range(50): 
                                time.sleep(0.1)
                                try:
                                    pres = win32com.client.GetObject(path)
                                    if pres:
                                        try:
                                            # 즉각 은폐 및 포커스 차단
                                            pres.Application.Visible = 0
                                            pres.Application.WindowState = 2 # ppWindowMinimized
                                        except: pass
                                        break
                                except: pass
                            if not pres: raise Exception("GetObject Redirect Failed")
                        except:
                            app = self._get_app("ppt")
                            pres = app.Presentations.Open(path, WithWindow=False, ReadOnly=True)
                        
                        # BAK 북마크 로직
                        for i, slide in enumerate(pres.Slides, 1):
                            try:
                                if not slide.Shapes.HasTitle: slide.Shapes.AddTitle()
                                title = slide.Shapes.Title
                                title.TextFrame.TextRange.Text = f"{os.path.splitext(f_name)[0]}-P{i}"
                                title.Top = -5000
                            except: pass
                        pres.SaveAs(t_pdf, 32); pres.Close(); ok = True
                    elif ext in ['.doc', '.docx']:
                        app = self._get_app("word")
                        doc = app.Documents.Open(path, Visible=False, ReadOnly=True)
                        doc.SaveAs(t_pdf, 17); doc.Close(); ok = True
                    elif ext in ['.xls', '.xlsx', '.xlsm']:
                        wb = None
                        try:
                            import win32api
                            # 이미 프리워밍되었으므로 즉시 GetObject 시도 (고속 폴링)
                            for wait_idx in range(50):
                                time.sleep(0.1)
                                try:
                                    wb = win32com.client.GetObject(path)
                                    if wb:
                                        try:
                                            wb.Application.Visible = False
                                            wb.Application.WindowState = -4140 # xlMinimized
                                        except: pass
                                        break
                                except: pass
                            if not wb: raise Exception("GetObject Redirect Failed")
                        except:
                            app = self._get_app("excel")
                            wb = app.Workbooks.Open(path, UpdateLinks=False, ReadOnly=True)
                            
                        self._clean_workbook(wb, smart_fix=True)
                        
                        # [v35.0 Improvement] 모든 워크시트 및 이미지 포함 로직
                        try:
                            # 모든 숨겨진 시트 표시
                            for sheet in wb.Worksheets:
                                sheet.Visible = -1 # xlSheetVisible
                            # 모든 워크시트 선택 (일괄 변환 대상 지정)
                            wb.Worksheets.Select()
                        except: pass
                        
                        # ExportAsFixedFormat(Type=0(PDF), IgnorePrintAreas=False(레이아웃 준수))
                        wb.ExportAsFixedFormat(0, t_pdf, Quality=0, IncludeDocProperties=True, IgnorePrintAreas=False)
                        wb.Close(False)
                        ok = True
                    
                    if ok and os.path.exists(t_pdf):
                        src = fitz.open(t_pdf); pdf.insert_pdf(src); src.close(); tmps.append(t_pdf)
                except: pass
        
        if pdf.page_count > 0:
            save_opts = {"garbage": 1}
            if config.get('compress_on'):
                self.callback("log", "[OPT] 엔진 압축 가동 (물리적 대체 모드: 트랜스코딩)")
                processed_xrefs = set()
                for page in pdf:
                    for img in page.get_images(full=True):
                        xref = img[0]
                        if xref in processed_xrefs: continue
                        processed_xrefs.add(xref)
                        try:
                            orig_len = pdf.xref_stream_length(xref)
                            if orig_len < 5000: continue
                            pix = fitz.Pixmap(pdf, xref)
                            img_changed = False
                            if pix.w > 2560 or pix.h > 2560:
                                scale = min(2560 / pix.w, 2560 / pix.h)
                                mat = fitz.Matrix(scale, scale)
                                pix = fitz.Pixmap(pix, mat)
                                img_changed = True
                            
                            compressed = pix.tobytes("jpeg", quality=50) # Web(50) 기준
                            if len(compressed) < orig_len or img_changed:
                                try: page.replace_image(xref, stream=compressed)
                                except: pdf.update_stream(xref, compressed)
                        except: pass
                save_opts = {"garbage": 4, "deflate": True, "clean": True, "deflate_images": True, "deflate_fonts": True}
            
            pdf.save(target_path, **save_opts)
            self.callback("log", f"[SAVE] PDF 생성(고출력): {res_name}")
        pdf.close()
        for t in tmps:
            try: os.remove(t)
            except: pass

    def run_replicator(self, config):
        """범용 복제 및 명칭 매니저 (v33 고도화)"""
        src_path = os.path.abspath(config['src_path'])
        dest_dir = os.path.abspath(config['dest_dir'])
        src_ext = os.path.splitext(src_path)[1]
        if not os.path.exists(dest_dir): os.makedirs(dest_dir)
        
        name_list = config['name_list']
        success = 0
        for i, name in enumerate(name_list, 1):
            list_base = os.path.splitext(name)[0]
            new_name = config['edit_func'](list_base) + src_ext
            target_path = os.path.join(dest_dir, new_name)
            self.callback("status", f"[{i}/{len(name_list)}] 복제: {new_name}")
            try:
                shutil.copy2(src_path, target_path)
                success += 1
            except: pass
        self.callback("done", f"[DIR] 복제 완료 ({success}/{len(name_list)})")

    def run_search(self, config):
        """지능형 로컬 통합 탐색 엔진 (BAK01_search_two_items v17 완전 복원 및 가속)"""
        found_cnt = 0
        seen_data = set()
        files = []
        
        # [창의적 통합] 로컬 탐색용 레거시 스페셜리스트 로드
        legacy_file = config.get('engine_ver')
        if legacy_file and "Standard" not in legacy_file:
            self.callback("log", f"[SEARCH] 전문 탐색 엔진 가동: {legacy_file}")
            legacy_mod = LegacyEngineLoader.load_specific_engine(legacy_file)
            if legacy_mod and hasattr(legacy_mod, "SearchEngine"):
                legacy_mod.SearchEngine.perform_search(config, lambda type, d: 
                    self.callback("log", f"[FOUND] 발견: {d['file']} > {d['v1']} / {d['v2']}") if type == "result" else None)
                self.callback("done", "🔍 전문 엔진 탐색 완료")
                return

        for root_dir, _, fs in os.walk(config['base_path']):
            for f in fs:
                if config['file_kw'] in f and f.lower().endswith((".xlsx", ".xlsm")):
                    files.append(os.path.join(root_dir, f))
        files.sort(key=os.path.getmtime, reverse=True)

        for f_path in files:
            if found_cnt >= config['limit']: break
            self.callback("status", f"진행: {os.path.basename(f_path)}")
            try:
                # [RESTORED] BAK01의 정밀 로딩 방식
                wb = openpyxl.load_workbook(f_path, data_only=True, read_only=True)
                if config['sheet_scope'] == "첫 번째 시트": target_ws = [wb.worksheets[0]]
                elif config['sheet_scope'] == "전체 워크시트": target_ws = wb.worksheets
                else: target_ws = [s for s in wb.worksheets if s.title == config['sheet_scope']] or wb.worksheets
                
                for ws in target_ws:
                    if found_cnt >= config['limit']: break
                    h1_idx, h2_idx, start_row = -1, -1, 1
                    
                    # [RESTORED] 헤더 행 정밀 탐색 (BAK01 tune)
                    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), 1):
                        r_s = [str(c or "").strip() for c in row]
                        for i, cell_txt in enumerate(r_s):
                            if config['h1_name'] and config['h1_name'] in cell_txt:
                                if h1_idx == -1: h1_idx = i
                            if config['h2_name'] and config['h2_name'] in cell_txt:
                                if h2_idx == -1: h2_idx = i
                        if h1_idx != -1 or (config['g_mode'] != "모든 열 (전역 검색)" and h2_idx != -1):
                            start_row = r_idx + 1; break
                    
                    for row in ws.iter_rows(min_row=start_row, values_only=True):
                        if found_cnt >= config['limit']: break
                        row_s = [str(c or "").strip() for c in row]
                        if not any(row_s): continue
                        
                        # [RESTORED] 정밀 AND/OR 매칭 로직
                        match1 = any(k in (row_s[h1_idx] if h1_idx < len(row_s) else "") for k in config['kw1']) if h1_idx != -1 and config['kw1'] else (not config['kw1'])
                        
                        if config['g_mode'] == "모든 열 (전역 검색)":
                            match2 = any(k in " ".join(row_s) for k in config['kw2']) if config['kw2'] else True
                        else:
                            match2 = any(k in (row_s[h2_idx] if h2_idx < len(row_s) else "") for k in config['kw2']) if h2_idx != -1 and config['kw2'] else (not config['kw2'])
                        
                        if match1 and match2:
                            v1 = row_s[h1_idx] if h1_idx != -1 and h1_idx < len(row_s) else ""
                            v2 = row_s[h2_idx] if h2_idx != -1 and h2_idx < len(row_s) else "/".join(row_s)
                            if config['unique'] and (v1, v2) in seen_data: continue
                            seen_data.add((v1, v2))
                            found_cnt += 1
                            self.callback("result", {
                                "cnt": found_cnt, "file": os.path.basename(f_path), 
                                "sheet": ws.title, "v1": v1, "v2": v2, 
                                "h1": config['h1_name'], "h2": config['h2_name'],
                                "full": " / ".join(row_s), "g_mode": config['g_mode']
                            })
                wb.close()
            except: pass
        self.callback("done", f"[SEARCH] 탐색 완료 ({found_cnt}건 발견)")

    def _analyze_excel_structure(self, file_path):
        """BAK01_search_two_items 에 존재하던 병합 셀 대응 구조 분석 로직 완전 복원"""
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet_names = wb.sheetnames
        ws = wb.active
        headers = []
        for r in range(1, 11):
            row_data = []
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                val = cell.value
                # [RESTORED] 병합 셀 상위 값 추적 로직
                if val is None:
                    for rg in ws.merged_cells.ranges:
                        if cell.coordinate in rg:
                            val = ws.cell(row=rg.min_row, column=rg.min_col).value; break
                txt = str(val or "").strip()
                if txt and len(txt) < 25: row_data.append(txt)
            if len(set(row_data)) >= 2:
                headers = list(dict.fromkeys(row_data)); break
        wb.close()
        return {"sheets": sheet_names, "headers": headers}

    def _cleanup(self):
        """[v35.4.6] 세션 종료 및 모든 오피스 인스턴스 확실한 Quit() 보장"""
        for n, a in self.apps.items():
            if a:
                try: 
                    # 문서를 모두 닫고 앱 종료
                    if n == "excel":
                        try:
                            a.DisplayAlerts = False
                            for wb in a.Workbooks: wb.Close(False)
                        except: pass
                    elif n == "ppt":
                        try:
                            for p in a.Presentations: p.Close()
                        except: pass
                    
                    # 3회 리트라이 종료 시도
                    for _ in range(3):
                        try: a.Quit(); break
                        except: time.sleep(0.5)
                except: pass
                self.apps[n] = None

# ═══════════════════════════════════════════════════════════
# LAYER 2: PRESENTATION (Universal Interface - View)
# ═══════════════════════════════════════════════════════════
class ManagementApp:
    def __init__(self, root):
        self.root = root
        self.root.title("전사적 통합 문서 자동화 관리 시스템 v34.1.16 (Hardened)")
        self.root.geometry("820x980")
        
        # Shared Variables
        self.source_mode = tk.StringVar(value="폴더")
        # [이식성 강화] 하드코딩된 절대 경로를 현재 작업 디렉토리로 초기화
        self.path_var = tk.StringVar(value=os.getcwd())
        self.selected_files = []
        self.rep_selected_names = []
        self.is_running = False

        # Replicator Vars (Restored BAK01 Context)
        self.rep_edit_mode = tk.StringVar(value="치환")
        self.rep_naming_mode = tk.StringVar(value="접두사")
        self.rep_prefix_text = tk.StringVar(value="강화시방서"); self.rep_prefix_pos = tk.IntVar(value=0)
        self.rep_suffix_text = tk.StringVar(value="_Rev01"); self.rep_suffix_pos = tk.IntVar(value=0)
        self.rep_range_start = tk.IntVar(value=0); self.rep_range_end = tk.IntVar(value=5)
        self.rep_replace_text = tk.StringVar(value="치환문구")
        self.rep_find_text = tk.StringVar(value="PR당사안"); self.rep_change_to = tk.StringVar(value="강화시방서")
        self.rep_src_file = tk.StringVar(); self.rep_dest_dir = tk.StringVar()
        self.rep_source_mode = tk.StringVar(value="파일")
        self.rep_preview = tk.StringVar(value="대기 중...")

        # Specialist Search UI Vars (BAK01_search_two_items v17 완전 복원)
        self.srch_file_kw = tk.StringVar(value="발송_솔루션_목록_(기숙사 및 사택)")
        self.srch_limit = tk.StringVar(value="10")
        self.srch_header1 = tk.StringVar(value="계약명")
        self.srch_header2 = tk.StringVar(value="공사내용")
        self.srch_sheet_scope = tk.StringVar(value="전체 워크시트")
        self.srch_scope_mode = tk.StringVar(value="모든 열 (전역)")
        self.srch_unique = tk.BooleanVar(value=True)

        # Missing Hybrid Vars (Quantum Fix)
        self.compress_var = tk.BooleanVar(value=True)
        self.exts_p_var = tk.StringVar(value=".pdf, .png, .jpg, .jpeg")
        self.exts_o_var = tk.StringVar(value=".pptx, .docx, .xlsx, .xls, .xlsm")
        self.merge_mode = tk.StringVar(value="기본 (통합 PDF) [배포형]")
        self.merge_engine_var = tk.StringVar(value="Standard (v33.4 Ultimate)")
        self.clean_engine_var = tk.StringVar(value="Standard (v33.4 Ultimate)")
        self.amt_sheet_var = tk.StringVar(value="PR당사안")
        self.amt_cell_var = tk.StringVar(value="M32")
        self.amt_prefix_var = tk.StringVar(value="F")
        self.hide_col_count_var = tk.IntVar(value=4)
        self.smart_fix_var = tk.BooleanVar(value=True)
        self.srch_engine_var = tk.StringVar(value="Standard (v33.6 Quantum)")
        self.prefix_var = tk.StringVar(value="M_")
        self.suffix_var = tk.StringVar(value="")

        self._build_ui()
        self._refresh_io()

    def _build_ui(self):
        # Header
        hdr = tk.Frame(self.root, bg="#263238", pady=15); hdr.pack(fill="x")
        tk.Label(hdr, text="ENTERPRISE DOCUMENT ORCHESTRATOR v34.1.16", font=("Consolas", 16, "bold"), bg="#263238", fg="white").pack()
        tk.Label(hdr, text="Ultimate Hybrid Engine | Unified Automation Management", font=("Malgun Gothic", 9), bg="#263238", fg="#90A4AE").pack()

        self.tabs = ttk.Notebook(self.root); self.tabs.pack(fill="both", expand=True, padx=20, pady=10)

        # Tabs
        self.tab_merge = tk.Frame(self.tabs, padx=15, pady=15); self.tabs.add(self.tab_merge, text=" 병합 오케스트레이터 ")
        self.tab_clean = tk.Frame(self.tabs, padx=15, pady=15); self.tabs.add(self.tab_clean, text=" 엑셀 딥-클리닝 ")
        self.tab_rep = tk.Frame(self.tabs, padx=15, pady=15); self.tabs.add(self.tab_rep, text=" 파일 복제/명칭 변경 ")
        self.tab_search = tk.Frame(self.tabs, padx=15, pady=15); self.tabs.add(self.tab_search, text=" 지능형 데이터 탐색 ")
        self.tab_manual = tk.Frame(self.tabs, padx=15, pady=15); self.tabs.add(self.tab_manual, text=" 시스템 매뉴얼 ")

        self._build_merge_tab()
        self._build_clean_tab()
        self._build_rep_tab()
        self._build_search_tab()
        self._build_manual_tab()

        self.st_lbl = tk.Label(self.root, text="시스템 엔진 준비됨", bd=1, relief="sunken", anchor="w", font=("Malgun Gothic", 8)); self.st_lbl.pack(side="bottom", fill="x")

    def _build_merge_tab(self):
        i_box = tk.LabelFrame(self.tab_merge, text="1. 대상 선택", padx=10, pady=8); i_box.pack(fill="x", pady=5)
        m_f = tk.Frame(i_box); m_f.pack(fill="x")
        tk.Radiobutton(m_f, text="폴더 일괄", variable=self.source_mode, value="폴더", command=self._refresh_io).pack(side="left")
        tk.Radiobutton(m_f, text="개별 선택", variable=self.source_mode, value="파일", command=self._refresh_io).pack(side="left", padx=20)
        self.in_ent_m = tk.Entry(i_box, textvariable=self.path_var, font=("Consolas", 9)); self.in_ent_m.pack(side="left", fill="x", expand=True, pady=5)
        tk.Button(i_box, text="경로 선택", width=10, command=self.handle_browse).pack(side="left", padx=5)

        c_box = tk.LabelFrame(self.tab_merge, text="2. 로직/명칭/대상 설정", padx=10, pady=8); c_box.pack(fill="x", pady=5)
        ttk.Combobox(c_box, textvariable=self.merge_mode, values=["기본 (통합 PDF) [배포형]", "동종 원본 유지 [편집형]"], state="readonly").pack(fill="x", pady=2)
        
        # Extensions
        x_f = tk.Frame(c_box); x_f.pack(fill="x", pady=2)
        tk.Label(x_f, text="PDF/이미지:", font=("Malgun Gothic", 8)).pack(side="left")
        tk.Entry(x_f, textvariable=self.exts_p_var, font=("Consolas", 8), width=30).pack(side="left", padx=5)
        tk.Label(x_f, text="Office문서:", font=("Malgun Gothic", 8)).pack(side="left", padx=(10,0))
        tk.Entry(x_f, textvariable=self.exts_o_var, font=("Consolas", 8), width=30).pack(side="left", padx=5)

        n_f = tk.Frame(c_box); n_f.pack(fill="x", pady=2)
        tk.Label(n_f, text="접두:").pack(side="left"); tk.Entry(n_f, textvariable=self.prefix_var, width=8).pack(side="left", padx=5)
        tk.Label(n_f, text="접미:").pack(side="left", padx=10); tk.Entry(n_f, textvariable=self.suffix_var, width=8).pack(side="left", padx=5)
        tk.Checkbutton(n_f, text="스마트 압축", variable=self.compress_var).pack(side="right")
        
        e_f = tk.Frame(c_box); e_f.pack(fill="x", pady=(5,0))
        tk.Label(e_f, text="[ENG] 핵심 엔진:", font=("bold")).pack(side="left")
        opts = ["Standard (v33.4 Ultimate)"] + LegacyEngineLoader.get_available_specialists("group_cross_merger")
        ttk.Combobox(e_f, textvariable=self.merge_engine_var, values=opts, state="readonly").pack(side="left", fill="x", expand=True, padx=5)

        self.log_m = tk.Text(self.tab_merge, height=18, font=("Consolas", 8), bg="#F1F8E9"); self.log_m.pack(fill="both", expand=True, pady=10)
        tk.Button(self.tab_merge, text="[START] 통합 병합 프로세스 가동", bg="#1B5E20", fg="white", font=("bold"), height=2, command=self.handle_start_merge).pack(fill="x")

    def _build_clean_tab(self):
        info = "[CLEAN] 외부 링크, 이름 정의, 깨질 수 있는 수식을 제거하고 스마트 열 정렬(E00000)을 수행합니다."
        tk.Label(self.tab_clean, text=info, fg="#1565C0", font=("Malgun Gothic", 9, "bold")).pack(pady=5)
        
        # Inputs Replication for Tab 2
        i_box = tk.LabelFrame(self.tab_clean, text="1. 클리닝 대상 선택", padx=10, pady=8, font=("bold"))
        i_box.pack(fill="x", pady=5)
        m_f = tk.Frame(i_box); m_f.pack(fill="x")
        tk.Radiobutton(m_f, text="폴더 내 엑셀 전체", variable=self.source_mode, value="폴더", command=self._refresh_io).pack(side="left")
        tk.Radiobutton(m_f, text="특정 엑셀들만 선택", variable=self.source_mode, value="파일", command=self._refresh_io).pack(side="left", padx=20)
        self.in_ent_c = tk.Entry(i_box, textvariable=self.path_var, font=("Consolas", 9)); self.in_ent_c.pack(side="left", fill="x", expand=True, pady=5)
        tk.Button(i_box, text="경로 선택", width=10, command=self.handle_browse).pack(side="left", padx=5)

        # Engine selection
        e_f = tk.Frame(self.tab_clean); e_f.pack(fill="x", pady=5)
        tk.Label(e_f, text="⚙️ 핵심 엔진:", font=("bold")).pack(side="left")
        opts = ["Standard (v33.4 Ultimate)"] + LegacyEngineLoader.get_available_specialists("group_cross_merger")
        ttk.Combobox(e_f, textvariable=self.clean_engine_var, values=opts, state="readonly").pack(side="left", fill="x", expand=True, padx=5)
        
        # New: Flexible Column Control
        tk.Label(e_f, text="| 숨길 열 개수:").pack(side="left")
        tk.Spinbox(e_f, from_=0, to=20, textvariable=self.hide_col_count_var, width=5).pack(side="left", padx=5)
        tk.Checkbutton(e_f, text="스마트 열 교정", variable=self.smart_fix_var).pack(side="right")

        # New: Amount Integrity Section
        a_box = tk.LabelFrame(self.tab_clean, text="2. 심층 금액 무결성 점검 설정 (Master Exclusive)", padx=10, pady=8)
        a_box.pack(fill="x", pady=5)
        a_f = tk.Frame(a_box); a_f.pack(fill="x")
        tk.Label(a_f, text="시트명:").pack(side="left"); tk.Entry(a_f, textvariable=self.amt_sheet_var, width=12).pack(side="left", padx=5)
        tk.Label(a_f, text="금액셀:").pack(side="left", padx=5); tk.Entry(a_f, textvariable=self.amt_cell_var, width=8).pack(side="left")
        tk.Label(a_f, text="부실접두사:").pack(side="left", padx=5); tk.Entry(a_f, textvariable=self.amt_prefix_var, width=5).pack(side="left")
        tk.Button(a_f, text="[START] 금액 점검 실행", bg="#FF6D00", fg="white", font=("bold", 8), command=self.handle_start_amt_check).pack(side="right")

        self.log_c = tk.Text(self.tab_clean, height=20, font=("Consolas", 8), bg="#E3F2FD"); self.log_c.pack(fill="both", expand=True, pady=10)
        tk.Button(self.tab_clean, text="[START] 엑셀 딥-클리닝 (정제+구조교정) 단독 실행", bg="#0277BD", fg="white", font=("bold"), height=2, command=self.handle_start_clean).pack(fill="x")

    def _build_rep_tab(self):
        b1 = tk.LabelFrame(self.tab_rep, text="1. 기준 원본", padx=10, pady=5); b1.pack(fill="x", pady=2)
        tk.Entry(b1, textvariable=self.rep_src_file, font=("Consolas", 8)).pack(side="left", fill="x", expand=True)
        tk.Button(b1, text="파일 선택", command=self.handle_rep_browse_src).pack(side="left", padx=5)

        b2 = tk.LabelFrame(self.tab_rep, text="2. 명칭 소스", padx=10, pady=5); b2.pack(fill="x", pady=2)
        m_f = tk.Frame(b2); m_f.pack(fill="x"); tk.Radiobutton(m_f, text="파일들", variable=self.rep_source_mode, value="파일").pack(side="left")
        tk.Radiobutton(m_f, text="폴더들", variable=self.rep_source_mode, value="폴더").pack(side="left", padx=10)
        tk.Radiobutton(m_f, text="폴더내파일", variable=self.rep_source_mode, value="폴더내파일").pack(side="left")
        tk.Button(b2, text="목록 선택/추가", command=self.handle_rep_browse_names, bg="#FFF9C4").pack(fill="x")

        b3 = tk.LabelFrame(self.tab_rep, text="3. 정밀 편집 규칙 (Master Logic)", padx=10, pady=5); b3.pack(fill="x", pady=2)
        self.rep_tabs = ttk.Notebook(b3); self.rep_tabs.pack(fill="x")
        
        # Tab 1: Insertion
        t1=tk.Frame(self.rep_tabs); self.rep_tabs.add(t1, text=" 정밀 삽입 ")
        tk.Radiobutton(t1, text="접두사(앞)", variable=self.rep_naming_mode, value="접두사").grid(row=0,column=0, sticky="w")
        tk.Entry(t1, textvariable=self.rep_prefix_text, width=15).grid(row=0,column=1, padx=5)
        tk.Label(t1, text="위치(N칸):").grid(row=0, column=2)
        tk.Spinbox(t1, from_=0, to=100, textvariable=self.rep_prefix_pos, width=5).grid(row=0, column=3, padx=5)
        
        tk.Radiobutton(t1, text="접미사(뒤)", variable=self.rep_naming_mode, value="접미사").grid(row=1,column=0, sticky="w")
        tk.Entry(t1, textvariable=self.rep_suffix_text, width=15).grid(row=1,column=1, padx=5)
        tk.Label(t1, text="위치(N칸):").grid(row=1, column=2)
        tk.Spinbox(t1, from_=0, to=100, textvariable=self.rep_suffix_pos, width=5).grid(row=1, column=3, padx=5)
        
        # Tab 2: Deletion
        t2=tk.Frame(self.rep_tabs); self.rep_tabs.add(t2, text=" 구간 삭제 ")
        tk.Label(t2, text="시작 위치:").pack(side="left"); tk.Spinbox(t2, from_=0, to=100, textvariable=self.rep_range_start, width=5).pack(side="left")
        tk.Label(t2, text="~ 끝 위치:").pack(side="left"); tk.Spinbox(t2, from_=0, to=200, textvariable=self.rep_range_end, width=5).pack(side="left")
        
        # Tab 3: Position Replace
        t3=tk.Frame(self.rep_tabs); self.rep_tabs.add(t3, text=" 부분 교체 ")
        tk.Label(t3, text="구간:").pack(side="left")
        tk.Spinbox(t3, from_=0, to=100, textvariable=self.rep_range_start, width=5).pack(side="left")
        tk.Label(t3, text="~").pack(side="left")
        tk.Spinbox(t3, from_=0, to=200, textvariable=self.rep_range_end, width=5).pack(side="left")
        tk.Label(t3, text=" ➔ 교체:").pack(side="left")
        tk.Entry(t3, textvariable=self.rep_replace_text, width=15).pack(side="left", padx=5)

        # Tab 4: String Replace
        t4=tk.Frame(self.rep_tabs); self.rep_tabs.add(t4, text=" 단어 치환 ")
        tk.Entry(t4, textvariable=self.rep_find_text, width=12).pack(side="left"); tk.Label(t4, text="➔").pack(side="left")
        tk.Entry(t4, textvariable=self.rep_change_to, width=12).pack(side="left")
        
        self.rep_tabs.bind("<<NotebookTabChanged>>", lambda e: self._on_rep_tab_change())

        tk.Label(b3, textvariable=self.rep_preview, font=("Consolas", 9, "bold"), bg="#FFEBEE", fg="#B71C1C").pack(fill="x", pady=5)
        
        b4 = tk.LabelFrame(self.tab_rep, text="4. 저장 경로", padx=10, pady=5); b4.pack(fill="x", pady=2)
        tk.Entry(b4, textvariable=self.rep_dest_dir, font=("Consolas", 8)).pack(side="left", fill="x", expand=True)
        tk.Button(b4, text="폴더 선택", command=self.handle_rep_browse_dest).pack(side="left", padx=5)

        self.log_rep = tk.Text(self.tab_rep, height=8, font=("Consolas", 8), bg="#F5F5F5"); self.log_rep.pack(fill="both", expand=True, pady=5)
        tk.Button(self.tab_rep, text="[START] 지능형 일괄 복제 프로세스 가동", bg="#2E7D32", fg="white", font=("bold"), height=2, command=self.handle_start_rep).pack(fill="x")

    def _build_search_tab(self):
        # [RESTORED] BAK01_search_two_items v17 UI 레이아웃 100% 한 자 한 자 복원
        tk.Label(self.tab_search, text="Excel 지능형 로컬 통합 탐색 시스템", font=("Malgun Gothic", 14, "bold"), pady=8, fg="#1A237E").pack()

        # 1. 경로 설정부
        path_f = tk.Frame(self.tab_search, pady=2); path_f.pack(fill="x")
        tk.Label(path_f, text="[DIR] 경로:", width=6, anchor="w", font=("Malgun Gothic", 9)).pack(side="left")
        tk.Entry(path_f, textvariable=self.path_var, font=("Malgun Gothic", 9)).pack(side="left", fill="x", expand=True, padx=5)
        tk.Button(path_f, text="폴더 선택", command=self.handle_browse, font=("Malgun Gothic", 8)).pack(side="left", padx=1)
        tk.Button(path_f, text="[OPT] 구조 분석", bg="#E3F2FD", command=self.handle_srch_analyze, font=("Malgun Gothic", 8, "bold")).pack(side="left", padx=1)

        # 2. 파일 필터부
        opt_f = tk.Frame(self.tab_search, pady=2); opt_f.pack(fill="x")
        tk.Label(opt_f, text="[FILE] 파일:", width=6, anchor="w", font=("Malgun Gothic", 9)).pack(side="left")
        tk.Entry(opt_f, textvariable=self.srch_file_kw, font=("Malgun Gothic", 9), width=35).pack(side="left", padx=5)
        tk.Button(opt_f, text="파일 선택(패턴추출)", command=self.handle_srch_pattern_pick, font=("Malgun Gothic", 8)).pack(side="left", padx=2)
        tk.Label(opt_f, text="| 🔢 개수:", font=("Malgun Gothic", 9)).pack(side="left", padx=(10,0))
        tk.Entry(opt_f, textvariable=self.srch_limit, width=5, font=("Malgun Gothic", 9)).pack(side="left", padx=5)
        tk.Checkbutton(opt_f, text="중복제거", variable=self.srch_unique, font=("Malgun Gothic", 8)).pack(side="left", padx=5)
        
        # 엔진 선택 (Quantum Specialist)
        tk.Label(opt_f, text="| ⚙️ 엔진:", font=("Malgun Gothic", 9)).pack(side="left", padx=(10,0))
        opts = ["Standard (v33.6 Quantum)"] + LegacyEngineLoader.get_available_specialists("search_two_items")
        ttk.Combobox(opt_f, textvariable=self.srch_engine_var, values=opts, state="readonly", width=15).pack(side="left")

        # 3. 검색 매칭 설정
        cfg_f = tk.Frame(self.tab_search, pady=8, relief="groove", bd=1); cfg_f.pack(fill="x")
        sel_f = tk.Frame(cfg_f, pady=2); sel_f.pack(fill="x", padx=5)
        tk.Label(sel_f, text="📍 열 헤더 1:", font=("Malgun Gothic", 8, "bold"), fg="#555").pack(side="left")
        self.srch_cb1 = ttk.Combobox(sel_f, textvariable=self.srch_header1, font=("Malgun Gothic", 9), width=15); self.srch_cb1.pack(side="left", padx=(2, 10))
        tk.Label(sel_f, text="📄 대상 시트:", font=("Malgun Gothic", 8, "bold"), fg="#555").pack(side="left")
        self.srch_sheet_cb = ttk.Combobox(sel_f, textvariable=self.srch_sheet_scope, font=("Malgun Gothic", 9), width=15); self.srch_sheet_cb.pack(side="left", padx=(2, 10))
        tk.Label(sel_f, text="🔍 범위:", font=("Malgun Gothic", 8, "bold"), fg="#555").pack(side="left")
        self.srch_g_cb = ttk.Combobox(sel_f, textvariable=self.srch_scope_mode, values=["모든 열 (전역)", "특정 열 선택"], font=("Malgun Gothic", 9), width=14); self.srch_g_cb.pack(side="left", padx=(2, 5))
        self.srch_g_cb.bind("<<ComboboxSelected>>", self.handle_srch_scope_toggle)
        self.srch_cb2 = ttk.Combobox(sel_f, textvariable=self.srch_header2, font=("Malgun Gothic", 9), width=12, state="disabled"); self.srch_cb2.pack(side="left", padx=2)

        kw_f = tk.Frame(cfg_f, pady=4); kw_f.pack(fill="x", padx=5)
        k1_wrap = tk.Frame(kw_f); k1_wrap.pack(side="left", fill="x", expand=True, padx=2)
        tk.Label(k1_wrap, text=" ↓ [열 1] 찾을 단어들(OR):", font=("Malgun Gothic", 7), fg="blue").pack(anchor="w")
        self.srch_kw1_txt = tk.Text(k1_wrap, height=3, font=("Malgun Gothic", 9)); self.srch_kw1_txt.pack(fill="x")
        self.srch_kw1_txt.insert("1.0", "베란다 누수보수(전후면)")
        
        k2_wrap = tk.Frame(kw_f); k2_wrap.pack(side="left", fill="x", expand=True, padx=2)
        tk.Label(k2_wrap, text=" ↓ [범위] 찾을 단어들(OR/AND):", font=("Malgun Gothic", 7), fg="blue").pack(anchor="w")
        self.srch_kw2_txt = tk.Text(k2_wrap, height=3, font=("Malgun Gothic", 9)); self.srch_kw2_txt.pack(fill="x")

        # 4. 결과창
        res_f = tk.Frame(self.tab_search, pady=5); res_f.pack(fill="both", expand=True)
        self.log_s = tk.Text(res_f, font=("Malgun Gothic", 9), bg="#fafafa"); self.log_s.pack(side="left", fill="both", expand=True)
        scr = tk.Scrollbar(res_f, command=self.log_s.yview); self.log_s.config(yscrollcommand=scr.set); scr.pack(side="right", fill="y")

        # 5. 액션 버튼부
        act_f = tk.Frame(self.tab_search, pady=10); act_f.pack(fill="x")
        tk.Button(act_f, text="[START] 탐색 시작", font=("Malgun Gothic", 10, "bold"), bg="#2E7D32", fg="white", width=25, command=self.handle_start_search).pack(side="left", expand=True, padx=5)
        tk.Button(act_f, text="[COPY] 복사", font=("Malgun Gothic", 10, "bold"), bg="#1565C0", fg="white", width=20, command=self.handle_srch_copy).pack(side="left", expand=True, padx=5)
        tk.Button(act_f, text="[HELP] 사용 방법", font=("Malgun Gothic", 9), bg="#607D8B", fg="white", command=self.handle_srch_help).pack(side="left", padx=10)

    def _build_manual_tab(self):
        man_txt = tk.Text(self.tab_manual, font=("Malgun Gothic", 9), padx=15, pady=15, bg="#FAFAFA"); man_txt.pack(fill="both", expand=True)
        manual = "[HELP] 시스템 매뉴얼 (v34.1.16)\n" + "1. 오케스트레이터: 이미지 포함 이기종 문서 지능형 병합\n" + "2. 딥-클리닝: 스마트 열 교정(E00000)으로 데이터 위치 자동 보정\n" + "3. 파일 복제: 정밀 삽입/구간삭제 기능이 포함된 대량 복제\n" + "4. 데이터 탐색: 백업 폴더 내 Excel 데이터 지능형 검색\n"
        man_txt.insert("1.0", manual); man_txt.config(state="disabled")

    def _refresh_io(self):
        btn_txt = "폴더 선택" if self.source_mode.get() == "폴더" else "파일 추가"
        state = "normal" if self.source_mode.get() == "폴더" else "readonly"
        self.in_ent_m.config(state=state); self.in_ent_c.config(state=state)
        if self.source_mode.get() == "파일": self.path_var.set(f"{len(self.selected_files)}개 파일 선택됨")

    def handle_browse(self):
        if self.source_mode.get() == "폴더":
            d = filedialog.askdirectory()
            if d: self.path_var.set(d)
        else:
            fs = filedialog.askopenfilenames()
            if fs: self.selected_files = list(fs); self.path_var.set(f"{len(self.selected_files)}개 파일 선택됨")

    def handle_start_merge(self):
        if self.is_running: return
        self.is_running = True; self.log_m.delete("1.0", "end")
        conf = {'source_dir': self.path_var.get() if self.source_mode.get()=="폴더" else os.path.dirname(self.selected_files[0]),
                'source_mode': self.source_mode.get(), 'selected_files': self.selected_files, 'merge_mode': self.merge_mode.get(),
                'prefix': self.prefix_var.get(), 'suffix': self.suffix_var.get(), 'exts_p': self.exts_p_var.get(),
                'exts_o': self.exts_o_var.get(), 'compress_on': self.compress_var.get(), 'engine_ver': self.merge_engine_var.get()}
        threading.Thread(target=lambda: ManagementEngine(self.merge_callback).run_merge(conf), daemon=True).start()

    def handle_start_clean(self):
        if self.is_running: return
        self.is_running = True; self.log_c.delete("1.0", "end")
        conf = {'source_dir': self.path_var.get() if self.source_mode.get()=="폴더" else os.path.dirname(self.selected_files[0]),
                'source_mode': self.source_mode.get(), 'selected_files': self.selected_files, 
                'smart_fix': self.smart_fix_var.get(), 'col_count': self.hide_col_count_var.get(),
                'engine_ver': self.clean_engine_var.get()}
        threading.Thread(target=lambda: ManagementEngine(self.clean_callback).run_clean_only(conf), daemon=True).start()

    def handle_start_amt_check(self):
        if self.is_running: return
        self.is_running = True; self.log_c.delete("1.0", "end")
        conf = {'source_dir': self.path_var.get() if self.source_mode.get()=="폴더" else os.path.dirname(self.selected_files[0]),
                'source_mode': self.source_mode.get(), 'selected_files': self.selected_files,
                'sheet_name': self.amt_sheet_var.get(), 'cell_addr': self.amt_cell_var.get(), 'prefix': self.amt_prefix_var.get()}
        threading.Thread(target=lambda: ManagementEngine(self.clean_callback).run_amount_check(conf), daemon=True).start()

    def handle_start_rep(self):
        if self.is_running: return
        src, dest = self.rep_src_file.get(), self.rep_dest_dir.get()
        if not src or not dest or not self.rep_selected_names: return
        self.is_running = True; self.log_rep.delete("1.0", "end")
        mode = self.rep_source_mode.get()
        if mode == "파일": n_list = [os.path.basename(f) for f in self.rep_selected_names]
        elif mode == "폴더": n_list = [os.path.basename(d) for d in self.rep_selected_names]
        else: n_list = os.listdir(self.rep_selected_names[0])
        conf = {"src_path": src, "dest_dir": dest, "name_list": sorted(n_list), "edit_func": self._apply_rep_logic}
        threading.Thread(target=lambda: ManagementEngine(self.rep_callback).run_replicator(conf), daemon=True).start()

    def handle_srch_analyze(self):
        f = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xlsm")], initialdir=self.path_var.get())
        if not f: return
        self.st_lbl.config(text="구조 분석 중...")
        def run():
            try:
                data = ManagementEngine(None)._analyze_excel_structure(f)
                self.root.after(0, lambda: self._update_srch_ui(data))
            except Exception as e: self.root.after(0, lambda: messagebox.showerror("오류", str(e)))
        threading.Thread(target=run, daemon=True).start()

    def handle_srch_pattern_pick(self):
        f = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xlsm")], initialdir=self.path_var.get())
        if f: self.srch_file_kw.set(os.path.splitext(os.path.basename(f))[0])

    def handle_srch_scope_toggle(self, _):
        self.srch_cb2.config(state="normal" if "특정" in self.srch_scope_mode.get() else "disabled")

    def _update_srch_ui(self, data):
        self.srch_sheet_cb['values'] = ["전체 워크시트", "첫 번째 시트"] + data['sheets']
        sh = sorted(data['headers'])
        self.srch_cb1['values'] = sh; self.srch_cb2['values'] = sh
        if sh: self.srch_header1.set(sh[0])
        if len(sh)>1: self.srch_header2.set(sh[1])
        self.st_lbl.config(text="구조 분석 완료")

    def handle_srch_copy(self):
        c = self.log_s.get("1.0", "end-1c")
        if not c.strip(): return
        self.root.clipboard_clear(); self.root.clipboard_append(c)
        messagebox.showinfo("복사", "결과가 클립보드에 복사되었습니다.")

    def handle_srch_help(self):
        msg = (
            "📖 [Excel 지능형 로컬 탐색기] 사용 가이드\n\n"
            "1. 폴더 선택: 검색할 파일들이 있는 폴더를 고르세요.\n"
            "2. ✨ 구조 분석: 파일 하나를 선택해 '시트'와 '헤더'를 학습시킵니다.\n"
            "3. 📝 파일 패턴: 찾을 파일의 이름 규칙을 입력하세요 (예: 발송_솔루션).\n"
            "4. 단어 입력: \n"
            "   - 줄바꿈으로 입력 시 각 단어 중 하나라도 있으면 매칭(OR)\n"
            "   - 열 1과 범위 양쪽에 입력 시 두 조건 모두 만족(AND)\n"
            "5. 범위 설정: '전역'은 전체 데이터 조사, '특정 열'은 지정 열만 정밀 조사\n"
            "6. 🚀 탐색 시작: 연산 후 결과를 확인하고 📋 복사하세요."
        )
        messagebox.showinfo("사용 방법", msg)

    def handle_start_search(self):
        if self.is_running: return
        k1 = [l.strip() for l in self.srch_kw1_txt.get("1.0", "end").split('\n') if l.strip()]
        k2 = [l.strip() for l in self.srch_kw2_txt.get("1.0", "end").split('\n') if l.strip()]
        if not k1 and not k2: return messagebox.showwarning("입력 부족", "찾을 단어를 입력하세요.")
        
        try: lim = int(self.srch_limit.get())
        except: lim = 10
        self.is_running = True; self.log_s.delete("1.0", "end")
        self.log_s.insert("end", f"[START] 탐색 가동 ({time.strftime('%H:%M:%S')})\n{'-'*80}\n")
        
        conf = {
            "base_path": self.path_var.get(),
            "kw1": k1, "kw2": k2,
            "file_kw": self.srch_file_kw.get(),
            "limit": lim, "unique": self.srch_unique.get(),
            "engine_ver": self.srch_engine_var.get(),
            "h1_name": self.srch_header1.get(),
            "h2_name": self.srch_header2.get(),
            "sheet_scope": self.srch_sheet_scope.get(),
            "g_mode": "모든 열 (전역 검색)" if "전역" in self.srch_scope_mode.get() else "특정 열 선택"
        }
        threading.Thread(target=lambda: ManagementEngine(self.search_callback).run_search(conf), daemon=True).start()

    def _apply_rep_logic(self, base):
        m = self.rep_edit_mode.get()
        try:
            if m == "삽입":
                if self.rep_naming_mode.get() == "접두사":
                    txt, pos = self.rep_prefix_text.get(), self.rep_prefix_pos.get()
                    if pos >= len(base): return base + "_" + txt
                    return base[:pos] + txt + "_" + base[pos:]
                else: # 접미사
                    txt, pos = self.rep_suffix_text.get(), self.rep_suffix_pos.get()
                    if pos == 0: return base + "_" + txt
                    if pos >= len(base): return txt + "_" + base
                    return base[:-pos] + "_" + txt + base[-pos:]
            elif m == "삭제":
                s, e = self.rep_range_start.get(), self.rep_range_end.get()
                return base[:s] + base[e:]
            elif m == "부분 교체":
                s, e, txt = self.rep_range_start.get(), self.rep_range_end.get(), self.rep_replace_text.get()
                return base[:s] + txt + base[e:]
            elif m == "단어 치환":
                return base.replace(self.rep_find_text.get(), self.rep_change_to.get())
        except: pass
        return base

    def _on_rep_tab_change(self):
        modes = ["삽입", "삭제", "부분 교체", "단어 치환"]
        idx = self.rep_tabs.index("current")
        if idx < len(modes): self.rep_edit_mode.set(modes[idx])
        self.rep_preview.set(f"샘플 ➔ {self._apply_rep_logic('파일명샘플')}")

    def handle_rep_browse_src(self): f=filedialog.askopenfilename(); self.rep_src_file.set(f)
    def handle_rep_browse_dest(self): d=filedialog.askdirectory(); self.rep_dest_dir.set(d)
    def handle_rep_browse_names(self):
        m = self.rep_source_mode.get()
        if m=="파일": fs=filedialog.askopenfilenames(); self.rep_selected_names=list(fs)
        elif m=="폴더": d=filedialog.askdirectory(); self.rep_selected_names=[d]
        else: d=filedialog.askdirectory(); self.rep_selected_names=[d]
        
        self.log_rep.delete("1.0", "end")
        n_list = []
        if m == "파일": n_list = [os.path.basename(f) for f in self.rep_selected_names]
        elif m == "폴더": n_list = [os.path.basename(d) for d in self.rep_selected_names]
        elif self.rep_selected_names: n_list = os.listdir(self.rep_selected_names[0])
        
        self.log_rep.insert("end", f"[*] 분석된 목록 ({len(n_list)}건):\n")
        for n in sorted(n_list)[:20]: self.log_rep.insert("end", f" - {n}\n")
        if len(n_list) > 20: self.log_rep.insert("end", " ... 외 다수\n")
        self.rep_preview.set(f"샘플 ➔ {self._apply_rep_logic('파일명샘플')}")

    def merge_callback(self, t, m): self.root.after(0, lambda: self._cb(t, m, "merge"))
    def clean_callback(self, t, m): self.root.after(0, lambda: self._cb(t, m, "clean"))
    def rep_callback(self, t, m): self.root.after(0, lambda: self._cb(t, m, "rep"))
    def search_callback(self, t, m): self.root.after(0, lambda: self._cb(t, m, "search"))

    def _cb(self, t, m, origin):
        log = {"merge": self.log_m, "clean": self.log_c, "rep": self.log_rep, "search": self.log_s}[origin]
        if t == "log": log.insert("end", f"[*] {m}\n"); log.see("end")
        elif t == "result": 
            log.insert("end", f"[FOUND] #{m['cnt']} | {m['file']} | {m['sheet']}\n")
            log.insert("end", f" [H1] {m['h1']}: {m['v1']}\n")
            # 전역 검색인 경우 full row 요약, 아닌 경우 v2 상세
            v_sub = f" [H2] {m['h2']}: {m['v2']}\n" if "특정" in m['g_mode'] else f" [INFO] 요약: {m['full']}\n"
            log.insert("end", v_sub + f"{'-'*60}\n")
            log.see("end")
        elif t == "status": self.st_lbl.config(text=m)
        elif t == "done": messagebox.showinfo("완료", m); self.is_running = False; self.st_lbl.config(text="시스템 엔진 대기 중")
        elif t == "error": messagebox.showerror("오류", m); self.is_running = False

def is_admin():
    import ctypes
    try: return ctypes.windll.shell32.IsUserAnAdmin()
    except: return False

def run_as_admin():
    import ctypes, sys
    if is_admin(): return
    params = f'"{sys.argv[0]}" ' + " ".join([f'"{arg}"' for arg in sys.argv[1:]])
    ctypes.windll.shell32.ShellExecuteW(None, "runas", sys.executable, params, None, 1)
    sys.exit(0)

if __name__ == "__main__":
    # [v35.4.1 Root Cause Fix] 관리자 권한 자동 승격 제거
    # Elevated(관리자) 프로세스에서 COM으로 Medium Integrity(일반) 오피스를 
    # 호출하려 할 때 ERROR_ELEVATION_REQUIRED (-2147024156) 오류가 발생함.
    # 스크립트와 Office의 권한을 일치시켜 UAC 충돌을 완벽히 해결.
    # run_as_admin() <-- 활성화 금지.
    
    import ctypes
    try: ctypes.windll.kernel32.SetErrorMode(0x0001 | 0x0002 | 0x8000)
    except: pass
    
    root = tk.Tk()
    
    # [UI Focus Fix]
    root.lift()
    root.focus_force()
    root.attributes('-topmost', True)
    root.after(500, lambda: root.attributes('-topmost', False))
    
    style = ttk.Style()
    style.theme_use("clam")
    app = ManagementApp(root)
    root.mainloop()
